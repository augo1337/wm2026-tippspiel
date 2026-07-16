#!/usr/bin/env python3
"""
fetch_scores.py – Holt WM 2026 Ergebnisse von football-data.org,
aktualisiert Ergebnisse.xlsx und ruft auswertung.py auf.

Lokale Nutzung:
    python fetch_scores.py             # nur lokal
    python fetch_scores.py --publish   # + Google Drive Update

Auf GitHub Actions läuft es automatisch mit --publish via Cron.

API Key: Umgebungsvariable FOOTBALL_DATA_API_KEY setzen,
         oder in dieser Datei direkt eintragen (Zeile API_KEY).
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import os
import subprocess
import requests
import openpyxl
from datetime import datetime, timezone, timedelta
from pathlib import Path

BASE         = Path(__file__).parent
ERGEBNISSE   = BASE / "Ergebnisse.xlsx"
TIPPS_ORDNER = BASE / "Tipps"

# ──────────────────────────────────────────────────────────────
# SPIELPLAN (Anstoß-Zeiten in UTC; Alle Zeiten CEST = UTC+2)
# Gruppenphase + KO bis Finale – wird für Zeitfenster-Check genutzt
# ──────────────────────────────────────────────────────────────
def _u(mo, d, h, mi=0):
    return datetime(2026, mo, d, h, mi, tzinfo=timezone.utc)

MATCH_KICKOFFS_UTC = [
    # Gruppenphase (11.06.–28.06.)
    _u(6,11,19),                                              # 11.06
    _u(6,12, 2), _u(6,12,19),                                # 12.06
    _u(6,13, 1), _u(6,13,19), _u(6,13,22),                  # 13.06
    _u(6,14, 1), _u(6,14, 4), _u(6,14,17), _u(6,14,20), _u(6,14,23),  # 14.06
    _u(6,15, 2), _u(6,15,16), _u(6,15,19), _u(6,15,22),    # 15.06
    _u(6,16, 1), _u(6,16,19), _u(6,16,22),                  # 16.06
    _u(6,17, 1), _u(6,17, 4), _u(6,17,17), _u(6,17,20), _u(6,17,23),  # 17.06
    _u(6,18, 2), _u(6,18,16), _u(6,18,19), _u(6,18,22),    # 18.06
    _u(6,19, 1), _u(6,19,19), _u(6,19,22),                  # 19.06
    _u(6,20, 0,30), _u(6,20, 3), _u(6,20,17), _u(6,20,20), # 20.06
    _u(6,21, 0), _u(6,21, 4), _u(6,21,16), _u(6,21,19), _u(6,21,22),  # 21.06
    _u(6,22, 1), _u(6,22,17), _u(6,22,21),                  # 22.06
    _u(6,23, 0), _u(6,23, 3), _u(6,23,17), _u(6,23,20), _u(6,23,23),  # 23.06
    _u(6,24, 2), _u(6,24,19), _u(6,24,22),                  # 24.06
    _u(6,25, 1), _u(6,25,20), _u(6,25,23),                  # 25.06
    _u(6,26, 2), _u(6,26,19),                                # 26.06
    _u(6,27, 0), _u(6,27, 3), _u(6,27,21), _u(6,27,23,30), # 27.06
    _u(6,28, 2),                                             # 28.06 letzte Gruppenspiele
    # KO-Runden (28.06.–19.07.)
    _u(6,28,19),                                             # S16 Tag 1
    _u(6,29,17), _u(6,29,20,30),                            # S16
    _u(6,30, 1), _u(6,30,17), _u(6,30,21),                  # S16
    _u(7, 1, 1), _u(7, 1,16), _u(7, 1,20),                  # S16
    _u(7, 2, 0), _u(7, 2,19), _u(7, 2,23),                  # S16
    _u(7, 3, 3), _u(7, 3,18), _u(7, 3,22),                  # S16
    _u(7, 4, 1,30), _u(7, 4,17), _u(7, 4,21),               # S16 Ende / S8 Beginn
    _u(7, 5,20),                                             # S8
    _u(7, 6, 0), _u(7, 6,19),                                # S8
    _u(7, 7, 0), _u(7, 7,16), _u(7, 7,20),                  # S8 Ende
    _u(7, 9,20),                                             # VF
    _u(7,10,19),                                             # VF
    _u(7,11,21),                                             # VF
    _u(7,12, 1),                                             # VF Ende
    _u(7,14,19), _u(7,15,19),                                # HF
    _u(7,18,19),                                             # Platz 3
    _u(7,19,19),                                             # Finale
]

# Zeitfenster: wie viele Minuten nach Anstoß darf ein Run noch als "relevant" gelten?
# 360 min = 90 min Spielzeit + 30 min Verlängerung + 30 min Elfmeter + 120 min API-Verzögerungspuffer
MATCH_WINDOW_MIN = 360

def is_match_window():
    """True wenn gerade ein WM-Spiel läuft oder erst kürzlich geendet hat."""
    now = datetime.now(timezone.utc)
    window = timedelta(minutes=MATCH_WINDOW_MIN)
    return any(kickoff <= now <= kickoff + window for kickoff in MATCH_KICKOFFS_UTC)

# ──────────────────────────────────────────────────────────────
# TEAM-NAMEN MAPPING  (ESPN Englisch → Deutsch)
# ──────────────────────────────────────────────────────────────
TEAM_MAP = {
    "Mexico":                       "Mexiko",
    "South Africa":                 "Südafrika",
    "Korea Republic":               "Südkorea",
    "South Korea":                  "Südkorea",
    "Czechia":                      "Tschechien",
    "Czech Republic":               "Tschechien",
    "Canada":                       "Kanada",
    "Bosnia and Herzegovina":       "Bosnien-Herzegowina",
    "Bosnia-Herzegovina":           "Bosnien-Herzegowina",
    "Qatar":                        "Katar",
    "Switzerland":                  "Schweiz",
    "Brazil":                       "Brasilien",
    "Morocco":                      "Marokko",
    "Haiti":                        "Haiti",
    "Scotland":                     "Schottland",
    "United States":                "USA",
    "USA":                          "USA",
    "Paraguay":                     "Paraguay",
    "Australia":                    "Australien",
    "Turkey":                       "Türkei",
    "Türkiye":                      "Türkei",
    "Germany":                      "Deutschland",
    "Curaçao":                      "Curaçao",
    "Curacao":                      "Curaçao",
    "Côte d'Ivoire":                "Elfenbeinküste",
    "Ivory Coast":                  "Elfenbeinküste",
    "Ecuador":                      "Ecuador",
    "Netherlands":                  "Niederlande",
    "Japan":                        "Japan",
    "Sweden":                       "Schweden",
    "Tunisia":                      "Tunesien",
    "Belgium":                      "Belgien",
    "Egypt":                        "Ägypten",
    "Iran":                         "Iran",
    "New Zealand":                  "Neuseeland",
    "Spain":                        "Spanien",
    "Cape Verde":                   "Kap Verde",
    "Cabo Verde":                   "Kap Verde",
    "Saudi Arabia":                 "Saudi-Arabien",
    "Uruguay":                      "Uruguay",
    "France":                       "Frankreich",
    "Senegal":                      "Senegal",
    "Iraq":                         "Irak",
    "Norway":                       "Norwegen",
    "Argentina":                    "Argentinien",
    "Algeria":                      "Algerien",
    "Austria":                      "Österreich",
    "Jordan":                       "Jordanien",
    "Portugal":                     "Portugal",
    "DR Congo":                     "DR Kongo",
    "Congo DR":                     "DR Kongo",
    "Democratic Republic of Congo": "DR Kongo",
    "Uzbekistan":                   "Usbekistan",
    "Colombia":                     "Kolumbien",
    "England":                      "England",
    "Croatia":                      "Kroatien",
    "Ghana":                        "Ghana",
    "Panama":                       "Panama",
}

def to_de(name):
    """Englischen ESPN-Namen → Deutschen Namen."""
    return TEAM_MAP.get(name, name)

# ──────────────────────────────────────────────────────────────
# SPIELPLAN  (muss mit auswertung.py übereinstimmen)
# ──────────────────────────────────────────────────────────────
sys.path.insert(0, str(BASE))
from auswertung import GRUPPENSPIELE

# ──────────────────────────────────────────────────────────────
# ESPN API ABRUF  (kostenlos, kein Key nötig, sofortige Ergebnisse)
# ──────────────────────────────────────────────────────────────
ESPN_URL = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard"

def fetch_matches():
    """Holt alle WM 2026 Spiele von der ESPN API."""
    r = requests.get(ESPN_URL,
                     params={"dates": "20260611-20260719", "limit": 200},
                     timeout=15)
    if r.status_code != 200:
        print(f"FEHLER: ESPN API Status {r.status_code}")
        return []
    events = r.json().get("events", [])
    if not events:
        print("HINWEIS: ESPN API liefert keine Spiele.")
    return events

# ──────────────────────────────────────────────────────────────
# ERGEBNISSE AUSWERTEN (ESPN Format)
# ESPN status.type.name:
#   STATUS_SCHEDULED  → noch nicht gespielt
#   STATUS_IN_PROGRESS → läuft
#   STATUS_FULL_TIME  → beendet (90 min)
#   STATUS_FINAL_AET  → nach Verlängerung
#   STATUS_FINAL_PEN  → nach Elfmeter
# ESPN note: "GROUP STAGE", "ROUND OF 32", "ROUND OF 16", etc.
# ──────────────────────────────────────────────────────────────
ESPN_KO_MAP = {
    "ROUND OF 32":   "S16",
    "ROUND OF 16":   "S8",
    "QUARTERFINALS": "VF",
    "SEMIFINALS":    "HF",
    "FINAL":         "F",
}
ESPN_FINISHED = {"STATUS_FULL_TIME", "STATUS_FINAL_AET", "STATUS_FINAL_PEN"}

def parse_matches(api_events):
    gs_lookup = {(m["heim"], m["gast"]): m["id"] for m in GRUPPENSPIELE}

    gs_results = {}
    ko_results = {k: [] for k in ["S16", "S8", "VF", "HF", "F", "WM"]}
    unmatched = []

    for event in api_events:
        comps = event.get("competitions", [{}])[0]
        status_name = comps.get("status", {}).get("type", {}).get("name", "")
        note = (comps.get("notes") or [{}])[0].get("headline", "").upper()
        teams = comps.get("competitors", [])
        if len(teams) != 2:
            continue

        # ESPN liefert home zuerst (teams[0]) wenn homeAway="home"
        home = next((t for t in teams if t.get("homeAway") == "home"), teams[0])
        away = next((t for t in teams if t.get("homeAway") == "away"), teams[1])
        heim_de = to_de(home["team"]["displayName"])
        gast_de = to_de(away["team"]["displayName"])

ESPN_LIVE = {"STATUS_IN_PROGRESS", "STATUS_HALFTIME"}

def parse_matches(api_events):
    gs_lookup = {(m["heim"], m["gast"]): m["id"] for m in GRUPPENSPIELE}

    gs_results = {}
    gs_live = {}   # laufende Spiele: match_id → "1:0 (67')"
    ko_results = {k: [] for k in ["S16", "S8", "VF", "HF", "F", "WM", "P3", "P3_participants"]}
    unmatched = []

    for event in api_events:
        comps = event.get("competitions", [{}])[0]
        status_name = comps.get("status", {}).get("type", {}).get("name", "")
        status_detail = comps.get("status", {}).get("displayClock", "")
        note = (comps.get("notes") or [{}])[0].get("headline", "").upper()
        teams = comps.get("competitors", [])
        if len(teams) != 2:
            continue

        home = next((t for t in teams if t.get("homeAway") == "home"), teams[0])
        away = next((t for t in teams if t.get("homeAway") == "away"), teams[1])
        heim_de = to_de(home["team"]["displayName"])
        gast_de = to_de(away["team"]["displayName"])

        h_score = home.get("score")
        a_score = away.get("score")

        # ── Live-Zwischenstand ──
        if status_name in ESPN_LIVE and h_score is not None and a_score is not None:
            match_id = gs_lookup.get((heim_de, gast_de))
            if match_id:
                clock = f" ({status_detail})" if status_detail else ""
                gs_live[match_id] = f"{int(h_score)}:{int(a_score)}{clock}"

        if status_name not in ESPN_FINISHED:
            continue

        if h_score is None or a_score is None:
            print(f"  Warnung: Kein Score für {heim_de} vs {gast_de}")
            continue

        # ── Gruppenphase oder KO? ──
        # ESPN liefert bei KO-Spielen note="" – deshalb zuerst prüfen ob es ein bekanntes Gruppenspiel ist
        match_id = gs_lookup.get((heim_de, gast_de))
        is_group = match_id is not None or "GROUP" in note

        if is_group:
            if match_id:
                gs_results[match_id] = f"{int(h_score)}:{int(a_score)}"
            else:
                unmatched.append(f"{heim_de} vs {gast_de}")
        else:
            # KO-Runde: Sieger ermitteln
            winner_obj = next((t for t in teams if t.get("winner") is True), None)
            if not winner_obj:
                # Fallback: höhere Score gewinnt (außer Elfmeter-Note)
                if h_score > a_score:
                    winner_obj = home
                elif a_score > h_score:
                    winner_obj = away
            winner = to_de(winner_obj["team"]["displayName"]) if winner_obj else None

            # Runde bestimmen: über note oder Datum
            runde = next((v for k, v in ESPN_KO_MAP.items() if k in note), None)
            if not runde:
                # Datum-basierte Erkennung (WM 2026 Spielplan)
                # S16 (Sechzehntelfinale): 28.06–03.07
                # S8  (Achtelfinale):      04.07–07.07
                # VF  (Viertelfinale):     09.07–12.07
                # HF  (Halbfinale):        14.07–15.07
                # F   (Finale):            19.07
                event_date = event.get("date", "")[:10]  # "2026-06-28"
                # Datum-basierte Erkennung (WM 2026 Spielplan)
                # S16 endet 04.07 morgens (Kolumbien 01:30Z), S8 beginnt 04.07 nachmittags (17:00Z)
                event_datetime = event.get("date", "")[:16]  # "2026-07-04T01:30"
                event_date = event_datetime[:10]
                event_hour = int(event_datetime[11:13]) if len(event_datetime) >= 13 else 12
                is_early_july4 = event_date == "2026-07-04" and event_hour < 12
                if ("2026-06-28" <= event_date <= "2026-07-03") or is_early_july4:
                    runde = "S16"
                elif "2026-07-04" <= event_date <= "2026-07-07":
                    runde = "S8"
                elif "2026-07-09" <= event_date <= "2026-07-12":
                    runde = "VF"
                elif "2026-07-14" <= event_date <= "2026-07-15":
                    runde = "HF"
                elif event_date == "2026-07-18":
                    runde = "P3"
                elif event_date >= "2026-07-19":
                    runde = "F"

            if runde and winner:
                if runde == "F":
                    ko_results["F"] = [heim_de, gast_de]
                    ko_results["WM"] = [winner]
                elif runde == "P3":
                    # Beide Teilnehmer und Sieger speichern
                    ko_results["P3"] = [heim_de, gast_de, winner]
                elif winner not in ko_results[runde]:
                    ko_results[runde].append(winner)

            # HF-Verlierer als P3-Teilnehmer sammeln (beide HF-Spiele müssen gespielt sein)
            if runde == "HF" and winner:
                loser = gast_de if winner == heim_de else heim_de
                if not ko_results.get("P3_participants"):
                    ko_results["P3_participants"] = []
                if loser not in ko_results["P3_participants"]:
                    ko_results["P3_participants"].append(loser)

    if unmatched:
        print(f"  Warnung: {len(unmatched)} Spiele nicht zugeordnet: {unmatched[:5]}")

    return gs_results, ko_results, gs_live

# ──────────────────────────────────────────────────────────────
# GQ BERECHNUNG aus tatsächlichen Ergebnissen
# Top 2 jeder Gruppe + 8 beste Dritte qualifizieren sich
# ──────────────────────────────────────────────────────────────
def _calc_actual_qualifiers(gs_results):
    import re as _re

    def parse_score(s):
        if not s:
            return None
        m = _re.match(r"^(\d+):(\d+)$", str(s).strip())
        return (int(m.group(1)), int(m.group(2))) if m else None

    def build_table(group_id):
        st = {}
        for m in GRUPPENSPIELE:
            if m["gruppe"] != group_id:
                continue
            for t in (m["heim"], m["gast"]):
                if t not in st:
                    st[t] = {"name": t, "pts": 0, "gd": 0, "gf": 0, "played": 0}
        for m in GRUPPENSPIELE:
            if m["gruppe"] != group_id:
                continue
            r = parse_score(gs_results.get(m["id"]))
            if r is None:
                continue
            hg, ag = r
            h, a = m["heim"], m["gast"]
            st[h]["played"] += 1; st[a]["played"] += 1
            st[h]["gf"] += hg; st[h]["gd"] += hg - ag
            st[a]["gf"] += ag; st[a]["gd"] += ag - hg
            if hg > ag:   st[h]["pts"] += 3
            elif hg < ag: st[a]["pts"] += 3
            else:         st[h]["pts"] += 1; st[a]["pts"] += 1
        return sorted(st.values(), key=lambda t: (-t["pts"], -t["gd"], -t["gf"], t["name"]))

    qualifiers = []
    thirds = []
    for g in ["A","B","C","D","E","F","G","H","I","J","K","L"]:
        table = build_table(g)
        # Nur auswerten wenn alle 3 Spiele der Gruppe gespielt
        games_played = sum(1 for m in GRUPPENSPIELE if m["gruppe"] == g and gs_results.get(m["id"]))
        if games_played < 3:
            continue  # Gruppe noch nicht vollständig
        if len(table) >= 1: qualifiers.append(table[0]["name"])
        if len(table) >= 2: qualifiers.append(table[1]["name"])
        if len(table) >= 3:
            t3 = table[2]
            thirds.append({"team": t3["name"], "pts": t3["pts"], "gd": t3["gd"], "gf": t3["gf"]})

    thirds.sort(key=lambda t: (-t["pts"], -t["gd"], -t["gf"], t["team"]))
    for t in thirds[:8]:
        qualifiers.append(t["team"])
    return qualifiers

# ──────────────────────────────────────────────────────────────
# ERGEBNISSE.XLSX SCHREIBEN
# ──────────────────────────────────────────────────────────────
def write_ergebnisse(gs_results, ko_results):
    # Vorhandene valide Ergebnisse laden und mit neuen mergen
    existing_gs = {}
    existing_ko = {k: [] for k in ["GQ", "S16", "S8", "VF", "HF", "F", "WM"]}
    if ERGEBNISSE.exists():
        try:
            wb_old = openpyxl.load_workbook(ERGEBNISSE)
            if "Gruppenphase" in wb_old.sheetnames:
                for row in wb_old["Gruppenphase"].iter_rows(min_row=2, values_only=True):
                    mid, _, _, _, score = row[0], row[1], row[2], row[3], row[4]
                    if mid and score:
                        import re as _re
                        if _re.match(r"^\d+:\d+$", str(score).strip()):
                            existing_gs[str(mid)] = str(score).strip()
            for runde in ["GQ", "S16", "S8", "VF", "HF", "F", "WM"]:
                if runde in wb_old.sheetnames:
                    teams = []
                    for row in wb_old[runde].iter_rows(min_row=2, values_only=True):
                        val = row[1] if len(row) > 1 else row[0]
                        if val:
                            teams.append(str(val).strip())
                    existing_ko[runde] = teams
        except Exception as e:
            print(f"  Hinweis: Konnte alte Ergebnisse nicht lesen: {e}")

    # Neue Ergebnisse über alte mergen (neue überschreiben alte)
    merged_gs = {**existing_gs, **gs_results}
    merged_ko = {}
    for runde in ["GQ", "S16", "S8", "VF", "HF", "P3", "P3_participants", "F", "WM"]:
        merged_ko[runde] = ko_results.get(runde) if ko_results.get(runde) else existing_ko.get(runde, [])

    # GQ automatisch berechnen aus den tatsächlichen Gruppenspiel-Ergebnissen
    merged_ko["GQ"] = _calc_actual_qualifiers(merged_gs)
    if merged_ko["GQ"]:
        print(f"  GQ-Teams berechnet: {len(merged_ko['GQ'])}/32")

    # Finalisten aus HF-Siegern ableiten wenn HF voll (2 Sieger bekannt)
    hf_teams = merged_ko.get("HF", [])
    if len(hf_teams) >= 2 and len(merged_ko.get("F", [])) < 2:
        merged_ko["F"] = hf_teams[:2]  # beide HF-Sieger = Finalisten
    elif len(hf_teams) == 1 and not merged_ko.get("F"):
        merged_ko["F"] = hf_teams[:]   # erster Finalist bekannt

    # P3-Teilnehmer aus HF-Verlierern ableiten
    # P3_participants = beide HF-Verlierer (für Anzeige)
    # P3 = nur der echte Sieger (erst nach Spiel um Platz 3)
    p3_parts = merged_ko.get("P3_participants", [])
    # P3 nur setzen wenn es ein echtes Ergebnis des P3-Spiels gibt (nicht aus HF ableiten)
    # merged_ko["P3"] bleibt wie vom parse_matches/existing_ko gesetzt

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gruppenphase"
    ws.append(["ID", "Heim", "Gast", "Datum", "Ergebnis"])
    for m in GRUPPENSPIELE:
        ws.append([m["id"], m["heim"], m["gast"], m["datum"],
                   merged_gs.get(m["id"], "")])

    for runde in ["GQ", "S16", "S8", "VF", "HF", "P3", "P3_participants", "F", "WM"]:
        ws2 = wb.create_sheet(runde)
        ws2.append(["Slot", "Team"])
        for i, team in enumerate(merged_ko.get(runde, []), 1):
            ws2.append([i, team])

    wb.save(ERGEBNISSE)

# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────
def main():
    print("=" * 50)
    print("  WM 2026 – Automatisches Score-Update")
    print("=" * 50)

    # Tipps-Ordner prüfen
    if not TIPPS_ORDNER.exists() or not list(TIPPS_ORDNER.glob("*.json")):
        print(f"\nHINWEIS: Keine JSON-Tipps in '{TIPPS_ORDNER}' gefunden.")
        print("Workflow wartet auf Tipp-Dateien – kein Fehler.")
        sys.exit(0)

    # Fehlende Ergebnisse ermitteln (aus vorhandener xlsx)
    existing_gs = {}
    if ERGEBNISSE.exists():
        try:
            import re as _re2
            _wb = openpyxl.load_workbook(ERGEBNISSE)
            if "Gruppenphase" in _wb.sheetnames:
                for _row in _wb["Gruppenphase"].iter_rows(min_row=2, values_only=True):
                    _mid, _score = _row[0], _row[4]
                    if _mid and _score and _re2.match(r"^\d+:\d+$", str(_score).strip()):
                        existing_gs[str(_mid)] = str(_score).strip()
        except Exception:
            pass
    missing_results = len(GRUPPENSPIELE) - len(existing_gs)

    # KO-Phase Check: solange nicht alle Runden bis zum Finale bekannt sind
    ko_phase_active = False
    if missing_results == 0 and ERGEBNISSE.exists():
        try:
            _wb3 = openpyxl.load_workbook(ERGEBNISSE)
            s16_count = sum(1 for r in _wb3["S16"].iter_rows(min_row=2, values_only=True) if r[1]) if "S16" in _wb3.sheetnames else 0
            s8_count  = sum(1 for r in _wb3["S8"].iter_rows(min_row=2, values_only=True)  if r[1]) if "S8"  in _wb3.sheetnames else 0
            vf_count  = sum(1 for r in _wb3["VF"].iter_rows(min_row=2, values_only=True)  if r[1]) if "VF"  in _wb3.sheetnames else 0
            hf_count  = sum(1 for r in _wb3["HF"].iter_rows(min_row=2, values_only=True)  if r[1]) if "HF"  in _wb3.sheetnames else 0
            wm_count  = sum(1 for r in _wb3["WM"].iter_rows(min_row=2, values_only=True)  if r[1]) if "WM"  in _wb3.sheetnames else 0
            ko_phase_active = s16_count < 16 or s8_count < 8 or vf_count < 4 or hf_count < 2 or wm_count < 1
            print(f"  KO-Stand: S16={s16_count}/16, S8={s8_count}/8, VF={vf_count}/4, HF={hf_count}/2, WM={wm_count}/1")
        except Exception:
            pass

    # API abrufen wenn: Spiel im aktiven Zeitfenster ODER fehlende Ergebnisse ODER KO-Phase läuft
    if not is_match_window() and missing_results == 0 and not ko_phase_active:
        print("\nKein WM-Spiel im aktiven Zeitfenster und alle Ergebnisse vorhanden – überspringe API-Abruf.")
        # GQ trotzdem neu berechnen falls noch nicht vorhanden
        try:
            _wb2 = openpyxl.load_workbook(ERGEBNISSE)
            _gq_missing = "GQ" not in _wb2.sheetnames or sum(1 for _ in _wb2["GQ"].iter_rows(min_row=2)) == 0
            if _gq_missing:
                print("  GQ fehlt noch – berechne aus vorhandenen Ergebnissen ...")
                gq_teams = _calc_actual_qualifiers(existing_gs)
                if gq_teams:
                    ws_gq = _wb2.create_sheet("GQ") if "GQ" not in _wb2.sheetnames else _wb2["GQ"]
                    if "GQ" in _wb2.sheetnames and _wb2["GQ"].max_row > 1:
                        pass  # bereits vorhanden
                    else:
                        ws_gq.append(["Slot", "Team"])
                        for i, t in enumerate(gq_teams, 1):
                            ws_gq.append([i, t])
                        _wb2.save(ERGEBNISSE)
                        print(f"  GQ-Teams gespeichert: {len(gq_teams)}/32")
        except Exception as e:
            print(f"  Hinweis GQ: {e}")
    else:
        if is_match_window():
            print("\nSpiel im aktiven Zeitfenster – hole Ergebnisse ...")
        else:
            print(f"\n{missing_results} fehlende Ergebnisse – hole Ergebnisse trotz inaktivem Zeitfenster ...")

        api_matches = fetch_matches()

        if not api_matches:
            print("Keine Spieldaten verfügbar – überspringe API-Update.")
        else:
            print(f"  {len(api_matches)} Spiele in der API gefunden")
            gs_results, ko_results, gs_live = parse_matches(api_matches)

            finished_gs = sum(1 for v in gs_results.values() if v)
            live_gs = len(gs_live)
            print(f"  Gruppenphase abgeschlossen: {finished_gs}/72")
            if live_gs:
                print(f"  Live laufend: {live_gs} Spiel(e): {list(gs_live.values())}")
            ko_info = {k: len(v) for k, v in ko_results.items() if v}
            if ko_info:
                print(f"  KO-Teams bekannt: {ko_info}")

            # Ergebnisse.xlsx aktualisieren
            write_ergebnisse(gs_results, ko_results)
            print(f"\n✓ Ergebnisse.xlsx aktualisiert")

            # Live-Scores als JSON speichern für auswertung.py
            import json as _json
            live_path = BASE / "live_scores.json"
            with open(live_path, "w", encoding="utf-8") as _f:
                _json.dump(gs_live, _f, ensure_ascii=False)

    # Live-Datei anlegen falls noch nicht vorhanden
    import json as _json
    live_path = BASE / "live_scores.json"
    if not live_path.exists():
        with open(live_path, "w") as _f:
            _json.dump({}, _f)

    # Rangliste immer neu erstellen (auch vor WM-Start, damit neue Tipps sichtbar werden)
    print("\nErstelle Rangliste ...")
    result = subprocess.run(
        [sys.executable, "-X", "utf8",
         str(BASE / "auswertung.py"),
         str(TIPPS_ORDNER),
         str(ERGEBNISSE)],
    )
    sys.exit(result.returncode)


if __name__ == "__main__":
    main()
