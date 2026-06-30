#!/usr/bin/env python3
"""
WM 2026 Tippspiel – Auswertung
===============================
Liest Tipp-Dateien (JSON vom tippzettel.html oder Ordner mit JSONs)
+ Ergebnisse.xlsx und erstellt Rangliste.xlsx + Rangliste.html.

Verwendung:
    python auswertung.py <tipps_ordner_oder_datei> <ergebnisse.xlsx>

Beispiele:
    python auswertung.py Tipps/        Ergebnisse.xlsx
"""

import sys
import re
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from zoneinfo import ZoneInfo
    BERLIN_TZ = ZoneInfo("Europe/Berlin")
except ImportError:
    BERLIN_TZ = timezone(timedelta(hours=2))  # CEST fallback

def now_berlin():
    return datetime.now(timezone.utc).astimezone(BERLIN_TZ)

# ============================================================
# SPIELPLAN
# ============================================================

ALLE_TEAMS = [
    "Mexiko", "Südafrika", "Südkorea", "Tschechien",
    "Kanada", "Bosnien-Herzegowina", "Katar", "Schweiz",
    "Brasilien", "Marokko", "Haiti", "Schottland",
    "USA", "Paraguay", "Australien", "Türkei",
    "Deutschland", "Curaçao", "Elfenbeinküste", "Ecuador",
    "Niederlande", "Japan", "Schweden", "Tunesien",
    "Belgien", "Ägypten", "Iran", "Neuseeland",
    "Spanien", "Kap Verde", "Saudi-Arabien", "Uruguay",
    "Frankreich", "Senegal", "Irak", "Norwegen",
    "Argentinien", "Algerien", "Österreich", "Jordanien",
    "Portugal", "DR Kongo", "Usbekistan", "Kolumbien",
    "England", "Kroatien", "Ghana", "Panama",
]

GRUPPENSPIELE = [
    {"id": "A1", "datum": "11.06.", "uhrzeit": "21:00", "heim": "Mexiko",         "gast": "Südafrika",           "gruppe": "A"},
    {"id": "A2", "datum": "12.06.", "uhrzeit": "04:00", "heim": "Südkorea",       "gast": "Tschechien",          "gruppe": "A"},
    {"id": "A3", "datum": "18.06.", "uhrzeit": "18:00", "heim": "Tschechien",     "gast": "Südafrika",           "gruppe": "A"},
    {"id": "A4", "datum": "19.06.", "uhrzeit": "03:00", "heim": "Mexiko",         "gast": "Südkorea",            "gruppe": "A"},
    {"id": "A5", "datum": "25.06.", "uhrzeit": "03:00", "heim": "Tschechien",     "gast": "Mexiko",              "gruppe": "A"},
    {"id": "A6", "datum": "25.06.", "uhrzeit": "03:00", "heim": "Südafrika",      "gast": "Südkorea",            "gruppe": "A"},
    {"id": "B1", "datum": "12.06.", "uhrzeit": "21:00", "heim": "Kanada",         "gast": "Bosnien-Herzegowina", "gruppe": "B"},
    {"id": "B2", "datum": "13.06.", "uhrzeit": "21:00", "heim": "Katar",          "gast": "Schweiz",             "gruppe": "B"},
    {"id": "B3", "datum": "18.06.", "uhrzeit": "21:00", "heim": "Schweiz",        "gast": "Bosnien-Herzegowina", "gruppe": "B"},
    {"id": "B4", "datum": "19.06.", "uhrzeit": "00:00", "heim": "Kanada",         "gast": "Katar",               "gruppe": "B"},
    {"id": "B5", "datum": "24.06.", "uhrzeit": "21:00", "heim": "Schweiz",        "gast": "Kanada",              "gruppe": "B"},
    {"id": "B6", "datum": "24.06.", "uhrzeit": "21:00", "heim": "Bosnien-Herzegowina", "gast": "Katar",          "gruppe": "B"},
    {"id": "C1", "datum": "14.06.", "uhrzeit": "00:00", "heim": "Brasilien",      "gast": "Marokko",             "gruppe": "C"},
    {"id": "C2", "datum": "14.06.", "uhrzeit": "03:00", "heim": "Haiti",          "gast": "Schottland",          "gruppe": "C"},
    {"id": "C3", "datum": "20.06.", "uhrzeit": "00:00", "heim": "Schottland",     "gast": "Marokko",             "gruppe": "C"},
    {"id": "C4", "datum": "20.06.", "uhrzeit": "02:30", "heim": "Brasilien",      "gast": "Haiti",               "gruppe": "C"},
    {"id": "C5", "datum": "25.06.", "uhrzeit": "00:00", "heim": "Schottland",     "gast": "Brasilien",           "gruppe": "C"},
    {"id": "C6", "datum": "25.06.", "uhrzeit": "00:00", "heim": "Marokko",        "gast": "Haiti",               "gruppe": "C"},
    {"id": "D1", "datum": "13.06.", "uhrzeit": "03:00", "heim": "USA",            "gast": "Paraguay",            "gruppe": "D"},
    {"id": "D2", "datum": "14.06.", "uhrzeit": "06:00", "heim": "Australien",     "gast": "Türkei",              "gruppe": "D"},
    {"id": "D3", "datum": "19.06.", "uhrzeit": "21:00", "heim": "USA",            "gast": "Australien",          "gruppe": "D"},
    {"id": "D4", "datum": "20.06.", "uhrzeit": "05:00", "heim": "Türkei",         "gast": "Paraguay",            "gruppe": "D"},
    {"id": "D5", "datum": "26.06.", "uhrzeit": "04:00", "heim": "Türkei",         "gast": "USA",                 "gruppe": "D"},
    {"id": "D6", "datum": "26.06.", "uhrzeit": "04:00", "heim": "Paraguay",       "gast": "Australien",          "gruppe": "D"},
    {"id": "E1", "datum": "14.06.", "uhrzeit": "19:00", "heim": "Deutschland",    "gast": "Curaçao",             "gruppe": "E"},
    {"id": "E2", "datum": "15.06.", "uhrzeit": "01:00", "heim": "Elfenbeinküste", "gast": "Ecuador",             "gruppe": "E"},
    {"id": "E3", "datum": "20.06.", "uhrzeit": "22:00", "heim": "Deutschland",    "gast": "Elfenbeinküste",      "gruppe": "E"},
    {"id": "E4", "datum": "21.06.", "uhrzeit": "02:00", "heim": "Ecuador",        "gast": "Curaçao",             "gruppe": "E"},
    {"id": "E5", "datum": "25.06.", "uhrzeit": "22:00", "heim": "Curaçao",        "gast": "Elfenbeinküste",      "gruppe": "E"},
    {"id": "E6", "datum": "25.06.", "uhrzeit": "22:00", "heim": "Ecuador",        "gast": "Deutschland",         "gruppe": "E"},
    {"id": "F1", "datum": "14.06.", "uhrzeit": "22:00", "heim": "Niederlande",    "gast": "Japan",               "gruppe": "F"},
    {"id": "F2", "datum": "15.06.", "uhrzeit": "04:00", "heim": "Schweden",       "gast": "Tunesien",            "gruppe": "F"},
    {"id": "F3", "datum": "20.06.", "uhrzeit": "19:00", "heim": "Niederlande",    "gast": "Schweden",            "gruppe": "F"},
    {"id": "F4", "datum": "21.06.", "uhrzeit": "06:00", "heim": "Tunesien",       "gast": "Japan",               "gruppe": "F"},
    {"id": "F5", "datum": "26.06.", "uhrzeit": "01:00", "heim": "Japan",          "gast": "Schweden",            "gruppe": "F"},
    {"id": "F6", "datum": "26.06.", "uhrzeit": "01:00", "heim": "Tunesien",       "gast": "Niederlande",         "gruppe": "F"},
    {"id": "G1", "datum": "15.06.", "uhrzeit": "21:00", "heim": "Belgien",        "gast": "Ägypten",             "gruppe": "G"},
    {"id": "G2", "datum": "16.06.", "uhrzeit": "03:00", "heim": "Iran",           "gast": "Neuseeland",          "gruppe": "G"},
    {"id": "G3", "datum": "21.06.", "uhrzeit": "21:00", "heim": "Belgien",        "gast": "Iran",                "gruppe": "G"},
    {"id": "G4", "datum": "22.06.", "uhrzeit": "03:00", "heim": "Neuseeland",     "gast": "Ägypten",             "gruppe": "G"},
    {"id": "G5", "datum": "27.06.", "uhrzeit": "05:00", "heim": "Ägypten",        "gast": "Iran",                "gruppe": "G"},
    {"id": "G6", "datum": "27.06.", "uhrzeit": "05:00", "heim": "Neuseeland",     "gast": "Belgien",             "gruppe": "G"},
    {"id": "H1", "datum": "15.06.", "uhrzeit": "18:00", "heim": "Spanien",        "gast": "Kap Verde",           "gruppe": "H"},
    {"id": "H2", "datum": "16.06.", "uhrzeit": "00:00", "heim": "Saudi-Arabien",  "gast": "Uruguay",             "gruppe": "H"},
    {"id": "H3", "datum": "21.06.", "uhrzeit": "18:00", "heim": "Spanien",        "gast": "Saudi-Arabien",       "gruppe": "H"},
    {"id": "H4", "datum": "22.06.", "uhrzeit": "00:00", "heim": "Uruguay",        "gast": "Kap Verde",           "gruppe": "H"},
    {"id": "H5", "datum": "27.06.", "uhrzeit": "02:00", "heim": "Kap Verde",      "gast": "Saudi-Arabien",       "gruppe": "H"},
    {"id": "H6", "datum": "27.06.", "uhrzeit": "02:00", "heim": "Uruguay",        "gast": "Spanien",             "gruppe": "H"},
    {"id": "I1", "datum": "16.06.", "uhrzeit": "21:00", "heim": "Frankreich",     "gast": "Senegal",             "gruppe": "I"},
    {"id": "I2", "datum": "17.06.", "uhrzeit": "00:00", "heim": "Irak",           "gast": "Norwegen",            "gruppe": "I"},
    {"id": "I3", "datum": "22.06.", "uhrzeit": "23:00", "heim": "Frankreich",     "gast": "Irak",                "gruppe": "I"},
    {"id": "I4", "datum": "23.06.", "uhrzeit": "02:00", "heim": "Norwegen",       "gast": "Senegal",             "gruppe": "I"},
    {"id": "I5", "datum": "26.06.", "uhrzeit": "21:00", "heim": "Norwegen",       "gast": "Frankreich",          "gruppe": "I"},
    {"id": "I6", "datum": "26.06.", "uhrzeit": "21:00", "heim": "Senegal",        "gast": "Irak",                "gruppe": "I"},
    {"id": "J1", "datum": "17.06.", "uhrzeit": "03:00", "heim": "Argentinien",    "gast": "Algerien",            "gruppe": "J"},
    {"id": "J2", "datum": "17.06.", "uhrzeit": "06:00", "heim": "Österreich",     "gast": "Jordanien",           "gruppe": "J"},
    {"id": "J3", "datum": "22.06.", "uhrzeit": "19:00", "heim": "Argentinien",    "gast": "Österreich",          "gruppe": "J"},
    {"id": "J4", "datum": "23.06.", "uhrzeit": "05:00", "heim": "Jordanien",      "gast": "Algerien",            "gruppe": "J"},
    {"id": "J5", "datum": "28.06.", "uhrzeit": "04:00", "heim": "Algerien",       "gast": "Österreich",          "gruppe": "J"},
    {"id": "J6", "datum": "28.06.", "uhrzeit": "04:00", "heim": "Jordanien",      "gast": "Argentinien",         "gruppe": "J"},
    {"id": "K1", "datum": "17.06.", "uhrzeit": "19:00", "heim": "Portugal",       "gast": "DR Kongo",            "gruppe": "K"},
    {"id": "K2", "datum": "18.06.", "uhrzeit": "04:00", "heim": "Usbekistan",     "gast": "Kolumbien",           "gruppe": "K"},
    {"id": "K3", "datum": "23.06.", "uhrzeit": "19:00", "heim": "Portugal",       "gast": "Usbekistan",          "gruppe": "K"},
    {"id": "K4", "datum": "24.06.", "uhrzeit": "04:00", "heim": "Kolumbien",      "gast": "DR Kongo",            "gruppe": "K"},
    {"id": "K5", "datum": "28.06.", "uhrzeit": "01:30", "heim": "Kolumbien",      "gast": "Portugal",            "gruppe": "K"},
    {"id": "K6", "datum": "28.06.", "uhrzeit": "01:30", "heim": "DR Kongo",       "gast": "Usbekistan",          "gruppe": "K"},
    {"id": "L1", "datum": "17.06.", "uhrzeit": "22:00", "heim": "England",        "gast": "Kroatien",            "gruppe": "L"},
    {"id": "L2", "datum": "18.06.", "uhrzeit": "01:00", "heim": "Ghana",          "gast": "Panama",              "gruppe": "L"},
    {"id": "L3", "datum": "23.06.", "uhrzeit": "22:00", "heim": "England",        "gast": "Ghana",               "gruppe": "L"},
    {"id": "L4", "datum": "24.06.", "uhrzeit": "01:00", "heim": "Panama",         "gast": "Kroatien",            "gruppe": "L"},
    {"id": "L5", "datum": "27.06.", "uhrzeit": "23:00", "heim": "Panama",         "gast": "England",             "gruppe": "L"},
    {"id": "L6", "datum": "27.06.", "uhrzeit": "23:00", "heim": "Kroatien",       "gast": "Ghana",               "gruppe": "L"},
]

# ============================================================
# SPALTENNAMEN  (müssen exakt mit den Forms-Fragen übereinstimmen)
# ============================================================

NAME_COL = "Name (Vorname Nachname)"

def match_col(m):
    return f"[Gr.{m['gruppe']}] {m['heim']} vs {m['gast']} ({m['datum']})"

KO_COLS = {
    "S16": [f"[S16] Weiterkommer {i}" for i in range(1, 17)],
    "S8":  [f"[S8] Weiterkommer {i}"  for i in range(1, 9)],
    "VF":  [f"[VF] Weiterkommer {i}"  for i in range(1, 5)],
    "HF":  [f"[HF] Weiterkommer {i}"  for i in range(1, 3)],
    "P3":  ["[P3] Dritter Platz"],
    "F":   ["[F] Finalist 1", "[F] Finalist 2"],
    "WM":  ["[WM] Weltmeister"],
}

KO_PUNKTE = {"S16": 5, "S8": 10, "VF": 15, "HF": 20}

# ============================================================
# PUNKTE BERECHNUNG
# ============================================================

def parse_score(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    m = re.match(r"^\s*(\d+)\s*[:–\-]\s*(\d+)\s*$", str(s).strip())
    return (int(m.group(1)), int(m.group(2))) if m else None

def calc_group_pts(tipp, result):
    t, r = parse_score(tipp), parse_score(result)
    if t is None or r is None:
        return 0
    th, tg = t
    rh, rg = r
    if th == rh and tg == rg:
        return 4
    if (th - tg) == (rh - rg):
        return 3
    if (th > tg) == (rh > rg) and not (th == tg and rh != rg) and not (th != tg and rh == rg):
        return 2
    if th == tg and rh == rg:
        return 2
    return 0

def norm(name):
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    return str(name).strip().lower()

def calc_ko_pts(pred_list, actual_set, pts_each):
    pred = {norm(t) for t in pred_list if norm(t)}
    actual = {norm(t) for t in actual_set if norm(t)}
    return len(pred & actual) * pts_each

def calc_finale_pts(pred_list, actual_set):
    pred = {norm(t) for t in pred_list if norm(t)}
    actual = {norm(t) for t in actual_set if norm(t)}
    return 20 if len(actual) == 2 and pred == actual else 0

def calc_wm_pts(pred, actual):
    return 25 if norm(pred) and norm(actual) and norm(pred) == norm(actual) else 0

def calc_p3_pts(pred, actual_list):
    """10 pts if pick is a P3 participant; +15 pts if pick is the winner (max 25)."""
    if not norm(pred) or not actual_list:
        return 0
    participants = {norm(t) for t in actual_list if norm(t)}
    if not participants:
        return 0
    p_norm = norm(pred)
    if p_norm not in participants:
        return 0
    pts = 10
    winner = norm(actual_list[0]) if actual_list else ""
    if winner and p_norm == winner:
        pts += 15
    return pts

def calc_qualifiers_from_predictions(detail_gruppe):
    """Returns a set of 32 team names predicted to qualify from the group stage,
    derived from the player's group score predictions (same logic as tippzettel.html)."""
    def build_table(group_id):
        st = {}
        for m in GRUPPENSPIELE:
            if m["gruppe"] != group_id:
                continue
            for t in (m["heim"], m["gast"]):
                if t not in st:
                    st[t] = {"name": t, "pts": 0, "gd": 0, "gf": 0}
        for m in GRUPPENSPIELE:
            if m["gruppe"] != group_id:
                continue
            d = detail_gruppe.get(m["id"], {})
            r = parse_score(d.get("tipp"))
            if r is None:
                continue
            hg, ag = r
            h, a = m["heim"], m["gast"]
            st[h]["gf"] += hg; st[h]["gd"] += hg - ag
            st[a]["gf"] += ag; st[a]["gd"] += ag - hg
            if hg > ag:   st[h]["pts"] += 3
            elif hg < ag: st[a]["pts"] += 3
            else:         st[h]["pts"] += 1; st[a]["pts"] += 1
        return sorted(st.values(), key=lambda t: (-t["pts"], -t["gd"], -t["gf"], t["name"]))

    qualifiers = set()
    thirds = []
    for g in ["A","B","C","D","E","F","G","H","I","J","K","L"]:
        table = build_table(g)
        if len(table) >= 1: qualifiers.add(table[0]["name"])
        if len(table) >= 2: qualifiers.add(table[1]["name"])
        if len(table) >= 3:
            t3 = table[2]
            thirds.append({"team": t3["name"], "pts": t3["pts"], "gd": t3["gd"], "gf": t3["gf"]})
    thirds.sort(key=lambda t: (-t["pts"], -t["gd"], -t["gf"], t["team"]))
    for t in thirds[:8]:
        qualifiers.add(t["team"])
    return qualifiers

def calc_gq_pts(pred_set, actual_list):
    actual = {norm(t) for t in actual_list if norm(t)}
    pred   = {norm(t) for t in pred_set   if norm(t)}
    return len(pred & actual) * 3

# ============================================================
# ERGEBNISSE EINLESEN
# ============================================================

def read_ergebnisse(path):
    wb = openpyxl.load_workbook(path)

    # Gruppenphase
    gs = wb["Gruppenphase"]
    results = {}
    for row in gs.iter_rows(min_row=2, values_only=True):
        match_id, _, _, _, score = row[0], row[1], row[2], row[3], row[4]
        if match_id and score:
            s = str(score).strip()
            if re.match(r"^\d+:\d+$", s):  # nur valide Ergebnisse (kein "None:None")
                results[str(match_id)] = s

    # KO-Runden
    ko_results = {}
    for runde in ["GQ", "S16", "S8", "VF", "HF", "P3", "F", "WM"]:
        if runde in wb.sheetnames:
            sheet = wb[runde]
            teams = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Column A = slot number, Column B = team name
                val = row[1] if len(row) > 1 else row[0]
                if val:
                    teams.append(str(val).strip())
            ko_results[runde] = teams

    return results, ko_results

# ============================================================
# JSON TIPPS EINLESEN  (neues Format: tippzettel.html)
# ============================================================

def json_to_row(data):
    row = {NAME_COL: data.get("name", "Unbekannt")}
    for key, val in data.get("gruppenphase", {}).items():
        row[key] = val
    for runde_id, cols in KO_COLS.items():
        if runde_id in ("WM", "P3"):
            row[cols[0]] = data.get(runde_id, "")
        else:
            teams = data.get(runde_id, [])
            if not isinstance(teams, list):
                teams = []
            for i, col in enumerate(cols):
                row[col] = teams[i] if i < len(teams) else ""
    return row

def read_json_tipps(path):
    p = Path(path)
    if p.is_dir():
        files = sorted(p.glob("*.json"))
        if not files:
            raise FileNotFoundError(f"Keine JSON-Dateien in {p} gefunden.")
        rows = [json_to_row(json.loads(f.read_text(encoding="utf-8"))) for f in files]
        print(f"  JSON-Dateien gelesen: {len(rows)}")
    else:
        rows = [json_to_row(json.loads(p.read_text(encoding="utf-8")))]
        print(f"  JSON-Datei gelesen: {p.name}")
    return pd.DataFrame(rows)

# ============================================================
# FORMS-RESPONSES EINLESEN
# ============================================================

def read_responses(path):
    df = pd.read_excel(path)
    # Microsoft Forms fügt Metaspalten vorne hinzu – Name-Spalte suchen
    if NAME_COL not in df.columns:
        print(f"WARNUNG: Spalte '{NAME_COL}' nicht gefunden.")
        print(f"Vorhandene Spalten: {list(df.columns[:10])}")
    return df

# ============================================================
# AUSWERTUNG
# ============================================================

def auswerten(df, gs_results, ko_results, gs_live=None):
    if gs_live is None:
        gs_live = {}
    rows = []
    for _, resp in df.iterrows():
        name = str(resp.get(NAME_COL, "Unbekannt")).strip()
        if not name or name == "nan":
            continue

        pts_gruppe = 0
        pts_gruppe_live = 0  # vorläufige Punkte aus laufenden Spielen
        detail_gruppe = {}
        for m in GRUPPENSPIELE:
            col = match_col(m)
            tipp = resp.get(col)
            result = gs_results.get(m["id"])
            live_score = gs_live.get(m["id"])  # z.B. "1:0 (67')"

            p = calc_group_pts(tipp, result)
            pts_gruppe += p

            # Vorläufige Live-Punkte (nur Tendenz/Torstand, kein Endstand)
            p_live = 0
            if not result and live_score:
                live_clean = live_score.split(" (")[0]  # "1:0" ohne "(67')"
                p_live = calc_group_pts(tipp, live_clean)
                pts_gruppe_live += p_live

            detail_gruppe[m["id"]] = {
                "tipp": tipp, "result": result, "punkte": p,
                "live_score": live_score, "punkte_live": p_live if (not result and live_score) else None,
                "heim": m["heim"], "gast": m["gast"], "datum": m["datum"]
            }

        pts_ko = {}
        for runde, pts_each in KO_PUNKTE.items():
            pred = [resp.get(c) for c in KO_COLS[runde]]
            actual = ko_results.get(runde, [])
            pts_ko[runde] = calc_ko_pts(pred, actual, pts_each)

        pts_ko["P3"] = calc_p3_pts(resp.get(KO_COLS["P3"][0]), ko_results.get("P3", []))

        pred_qualifiers = calc_qualifiers_from_predictions(detail_gruppe)
        pts_ko["GQ"] = calc_gq_pts(pred_qualifiers, ko_results.get("GQ", []))

        pts_ko["WM"] = calc_wm_pts(resp.get(KO_COLS["WM"][0]), ko_results.get("WM", [None])[0] if ko_results.get("WM") else None)

        total = pts_gruppe + sum(pts_ko.values())
        total_live = total + pts_gruppe_live  # Gesamt inkl. laufende Spiele
        rows.append({
            "Name": name,
            "Gesamt": total,
            "GesamtLive": total_live,
            "PtsLive": pts_gruppe_live,
            "Gruppenphase": pts_gruppe,
            "GQ":  pts_ko.get("GQ", 0),
            "S16": pts_ko.get("S16", 0),
            "S8":  pts_ko.get("S8",  0),
            "VF":  pts_ko.get("VF",  0),
            "HF":  pts_ko.get("HF",  0),
            "P3":  pts_ko.get("P3",  0),
            "Finale": pts_ko.get("HF", 0),
            "Weltmeister": pts_ko.get("WM", 0),
            "_detail": detail_gruppe,
            "_resp": resp,
            "_gq_pred": sorted([t for t in pred_qualifiers]),
        })

    # Sortierung: nach GesamtLive (inkl. laufende Spiele) während Live, sonst Gesamt
    has_live = any(r["PtsLive"] > 0 for r in rows)
    rows.sort(key=lambda x: x["GesamtLive"] if has_live else x["Gesamt"], reverse=True)
    for i, r in enumerate(rows):
        r["Platz"] = i + 1
    return rows

# ============================================================
# EXCEL OUTPUT
# ============================================================

# Farben
CLR_HEADER   = "1A3A5C"   # Dunkelblau
CLR_GOLD     = "FFD700"
CLR_SILVER   = "C0C0C0"
CLR_BRONZE   = "CD7F32"
CLR_GREEN    = "D4EDDA"
CLR_YELLOW   = "FFF3CD"
CLR_RED      = "F8D7DA"
CLR_LIGHT    = "F0F4F8"
CLR_WHITE    = "FFFFFF"

def header_font():    return Font(bold=True, color="FFFFFF", name="Calibri", size=11)
def cell_font():      return Font(name="Calibri", size=10)
def bold_font():      return Font(bold=True, name="Calibri", size=10)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def write_rangliste(wb, rows):
    ws = wb.create_sheet("Rangliste")
    ws.sheet_view.showGridLines = False

    headers = ["Platz", "Name", "Gesamt", "Gruppenphase",
               "Gr.-Qualifikation", "Achtelfinale", "Viertelfinale", "Halbfinale",
               "Finalisten (+20)", "Platz 3", "Weltmeister"]
    col_widths = [8, 22, 10, 14, 18, 14, 14, 12, 14, 10, 14]

    # Header
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font()
        cell.fill = fill(CLR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    # Daten
    for r in rows:
        row_num = r["Platz"] + 1
        values = [r["Platz"], r["Name"], r["Gesamt"], r["Gruppenphase"],
                  r.get("GQ", 0), r["S16"], r["S8"], r["VF"], r.get("HF", 0), r.get("P3", 0), r["Weltmeister"]]

        platz_fill = CLR_LIGHT
        if r["Platz"] == 1: platz_fill = CLR_GOLD
        elif r["Platz"] == 2: platz_fill = CLR_SILVER
        elif r["Platz"] == 3: platz_fill = CLR_BRONZE

        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.font = bold_font() if c <= 3 else cell_font()
            cell.fill = fill(platz_fill if c <= 3 else (CLR_LIGHT if row_num % 2 == 0 else CLR_WHITE))
            cell.alignment = Alignment(horizontal="center" if c != 2 else "left", vertical="center")
            cell.border = thin_border()

    ws.freeze_panes = "A2"

def write_details(wb, rows, gs_results):
    ws = wb.create_sheet("Details Gruppenphase")
    ws.sheet_view.showGridLines = False

    # Kopfzeile
    ws.cell(1, 1, "Spiel").font = header_font()
    ws.cell(1, 1).fill = fill(CLR_HEADER)
    ws.cell(1, 1).border = thin_border()
    ws.column_dimensions["A"].width = 30

    ws.cell(1, 2, "Ergebnis").font = header_font()
    ws.cell(1, 2).fill = fill(CLR_HEADER)
    ws.cell(1, 2).border = thin_border()
    ws.column_dimensions["B"].width = 12

    for i, r in enumerate(rows):
        col = i + 3
        ws.cell(1, col, r["Name"]).font = header_font()
        ws.cell(1, col).fill = fill(CLR_HEADER)
        ws.cell(1, col).alignment = Alignment(horizontal="center")
        ws.cell(1, col).border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 14

    gruppe = None
    data_row = 2
    for m in GRUPPENSPIELE:
        if m["gruppe"] != gruppe:
            gruppe = m["gruppe"]
            ws.cell(data_row, 1, f"── Gruppe {gruppe} ──").font = Font(bold=True, name="Calibri", size=10, color=CLR_HEADER)
            ws.cell(data_row, 1).fill = fill("E8EFF7")
            ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=len(rows)+2)
            data_row += 1

        label = f"{m['heim']} vs {m['gast']} ({m['datum']})"
        ws.cell(data_row, 1, label).font = cell_font()
        ws.cell(data_row, 1).border = thin_border()

        actual = gs_results.get(m["id"], "–")
        ws.cell(data_row, 2, actual).font = bold_font()
        ws.cell(data_row, 2).alignment = Alignment(horizontal="center")
        ws.cell(data_row, 2).border = thin_border()

        for i, r in enumerate(rows):
            col = i + 3
            detail = r["_detail"].get(m["id"], {})
            tipp = detail.get("tipp", "–") or "–"
            pts  = detail.get("punkte", 0)
            cell_val = f"{tipp}  (+{pts})" if actual != "–" else str(tipp)
            c = ws.cell(data_row, col, cell_val)
            c.font = cell_font()
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()
            if actual != "–":
                if pts == 4:   c.fill = fill(CLR_GREEN)
                elif pts == 3: c.fill = fill("C8E6C9")
                elif pts == 2: c.fill = fill(CLR_YELLOW)
                else:          c.fill = fill(CLR_RED)

        data_row += 1

    ws.freeze_panes = "C2"

def write_ko_details(wb, rows, ko_results):
    ws = wb.create_sheet("Details KO-Runden")
    ws.sheet_view.showGridLines = False

    runden_info = [
        ("S16", "Achtelfinale-Teilnehmer (16)",   16, 5),
        ("S8",  "Viertelfinale-Teilnehmer (8)",    8, 10),
        ("VF",  "Halbfinale-Teilnehmer (4)",        4, 15),
        ("HF",  "Finalisten (2)",                   2, 20),
        ("WM",  "Weltmeister",                      1, 25),
    ]

    ws.cell(1, 1, "KO-Runde / Slot").font = header_font()
    ws.cell(1, 1).fill = fill(CLR_HEADER)
    ws.cell(1, 1).border = thin_border()
    ws.column_dimensions["A"].width = 32

    ws.cell(1, 2, "Tatsächlich").font = header_font()
    ws.cell(1, 2).fill = fill(CLR_HEADER)
    ws.cell(1, 2).border = thin_border()
    ws.column_dimensions["B"].width = 18

    for i, r in enumerate(rows):
        col = i + 3
        ws.cell(1, col, r["Name"]).font = header_font()
        ws.cell(1, col).fill = fill(CLR_HEADER)
        ws.cell(1, col).alignment = Alignment(horizontal="center")
        ws.cell(1, col).border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 16

    data_row = 2
    for runde_id, runde_label, slots, pts in runden_info:
        # Sektionszeile
        ws.cell(data_row, 1, f"── {runde_label} (je {pts} Pkt.) ──").font = Font(bold=True, name="Calibri", size=10, color=CLR_HEADER)
        ws.cell(data_row, 1).fill = fill("E8EFF7")
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=len(rows)+2)
        data_row += 1

        actual_teams = {norm(t) for t in ko_results.get(runde_id, []) if norm(t)}

        for i, r in enumerate(rows):
            resp = r["_resp"]
            cols = KO_COLS[runde_id]
            pred = [resp.get(c) for c in cols]

            # Tipps als komma-getrennte Liste in einer Zeile pro Teilnehmer
            # Schreibe nur eine Übersichtszeile pro Runde (alle Tipps zusammen)
            pass

        # Eine Zeile pro Slot
        for slot_idx in range(slots):
            ws.cell(data_row, 1, f"  Weiterkommer {slot_idx + 1}").font = cell_font()
            ws.cell(data_row, 1).border = thin_border()

            actual_at_slot = list(ko_results.get(runde_id, []))
            actual_val = actual_at_slot[slot_idx] if slot_idx < len(actual_at_slot) else "–"
            ws.cell(data_row, 2, actual_val).font = bold_font()
            ws.cell(data_row, 2).alignment = Alignment(horizontal="center")
            ws.cell(data_row, 2).border = thin_border()

            for i, r in enumerate(rows):
                resp = r["_resp"]
                col_key = KO_COLS[runde_id][slot_idx] if slot_idx < len(KO_COLS[runde_id]) else None
                pred_val = resp.get(col_key) if col_key else None
                pred_norm = norm(pred_val)
                is_correct = pred_norm and pred_norm in actual_teams

                c = ws.cell(data_row, i + 3, str(pred_val) if pred_val and not pd.isna(pred_val) else "–")
                c.font = cell_font()
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border()
                if actual_teams:
                    c.fill = fill(CLR_GREEN if is_correct else CLR_RED)

            data_row += 1

    ws.freeze_panes = "C2"

# ============================================================
# HTML RANGLISTE  – interaktiv mit Spielerfilter & Detailansicht
# ============================================================

def write_html_rangliste(rows, gs_results, ko_results, out_path, gs_live=None):
    import json as _json
    if gs_live is None:
        gs_live = {}

    def _clean(val):
        if val is None:
            return ""
        try:
            if pd.isna(val):
                return ""
        except (TypeError, ValueError):
            pass
        return str(val).strip()

    # ── Upcoming-Match-Erkennung ──────────────────────────
    _now  = now_berlin().replace(tzinfo=None)
    _soon = (_now + timedelta(days=2)).date()
    def _is_kommend(datum_str, has_result):
        if has_result:
            return False
        try:
            return datetime.strptime(datum_str.rstrip('.') + '.2026', '%d.%m.%Y').date() <= _soon
        except Exception:
            return False

    # ── Daten für JavaScript aufbereiten ──────────────────
    players_js = []
    for r in rows:
        spiele = []
        for m in GRUPPENSPIELE:
            d = r["_detail"].get(m["id"], {})
            ergebnis = _clean(d.get("result"))
            tipp     = _clean(d.get("tipp"))
            pts      = d.get("punkte", 0) if ergebnis else None
            live     = gs_live.get(m["id"], "")  # z.B. "1:0 (67')"
            d_live   = r["_detail"].get(m["id"], {})
            pts_live = d_live.get("punkte_live")  # vorläufige Punkte (None wenn nicht live)
            spiele.append({
                "id": m["id"], "gr": m["gruppe"],
                "heim": m["heim"], "gast": m["gast"], "datum": m["datum"], "uhrzeit": m.get("uhrzeit", ""),
                "tipp": tipp, "ergebnis": ergebnis, "punkte": pts,
                "live": live, "punkte_live": pts_live,
                "kommend": _is_kommend(m["datum"], bool(ergebnis) or bool(live)),
            })

        ko_tipps = {}
        for runde, cols in KO_COLS.items():
            if runde in ("WM", "P3"):
                ko_tipps[runde] = _clean(r["_resp"].get(cols[0]))
            else:
                ko_tipps[runde] = [_clean(r["_resp"].get(c)) for c in cols]

        players_js.append({
            "platz": r["Platz"], "name": r["Name"],
            "gesamt": r["Gesamt"], "gesamt_live": r.get("GesamtLive", r["Gesamt"]),
            "pts_live": r.get("PtsLive", 0),
            "gruppe": r["Gruppenphase"],
            "gq": r.get("GQ", 0),
            "s16": r["S16"], "s8": r["S8"],
            "vf": r["VF"], "hf": r.get("HF", 0),
            "p3": r.get("P3", 0),
            "finale": r["Finale"], "wm": r["Weltmeister"],
            "spiele": spiele, "ko_tipps": ko_tipps,
            "gq_teams": r.get("_gq_pred", []),
        })

    ko_res_js = {
        k: [str(t).strip().lower() for t in v if t and str(t).strip()]
        for k, v in ko_results.items()
    }
    # Original-Schreibweise separat für Anzeige (nicht lowercase)
    ko_res_display = {
        k: [str(t).strip() for t in v if t and str(t).strip()]
        for k, v in ko_results.items()
    }
    # P3-Teilnehmer = HF-Verlierer (alle HF-Teams minus Finalisten)
    hf_teams = [str(t).strip() for t in ko_results.get("HF", []) if t and str(t).strip()]
    f_teams  = [str(t).strip() for t in ko_results.get("F",  []) if t and str(t).strip()]
    p3_participants = [t for t in hf_teams if t not in f_teams][:2]
    ko_res_display["P3_participants"] = p3_participants

    filled  = sum(1 for v in gs_results.values() if v)
    ts      = now_berlin().strftime("%d.%m.%Y %H:%M Uhr")
    leader  = rows[0]["Name"] if rows else "–"
    lpts    = rows[0]["Gesamt"] if rows else 0
    n       = len(rows)
    data_json = _json.dumps({"players": players_js, "ko": ko_res_js, "ko_display": ko_res_display, "live": gs_live}, ensure_ascii=False)

    # ── History laden + aktualisieren ────────────────────────
    import re as _re, calendar as _cal
    _history = []
    if Path(out_path).exists():
        try:
            _old_html = Path(out_path).read_text(encoding='utf-8')
            _hm = _re.search(r'const HISTORY=(\[.*?\]);', _old_html, _re.DOTALL)
            if _hm:
                _history = _json.loads(_hm.group(1))
        except Exception:
            _history = []
    _curr_pts = {r["Name"]: r["Gesamt"] for r in rows}
    if not _history or _history[-1].get("pts") != _curr_pts:
        _history.append({"ts": now_berlin().strftime("%d.%m %H:%M"), "pts": _curr_pts})
    _history = _history[-120:]
    history_json = _json.dumps(_history, ensure_ascii=False)

    # ── Live-Spielplan für JS-Banner ──────────────────────────
    _live_sched = []
    for _lm in GRUPPENSPIELE:
        try:
            _ld, _lmo = (int(x) for x in _lm["datum"].rstrip('.').split('.')[:2])
            _lhh, _lmi = (int(x) for x in (_lm.get("uhrzeit") or "21:00").split(':'))
            _l_utc = datetime(2026, _lmo, _ld, _lhh, _lmi) - timedelta(hours=2)
            _live_sched.append([int(_cal.timegm(_l_utc.timetuple())) * 1000,
                                 f"{_lm['heim']} vs {_lm['gast']}", 120])
        except Exception:
            pass
    for (_lmo2, _ld2, _lh2, _lmi2) in [
        (6,28,19,0),(6,29,17,0),(6,29,20,30),(6,30,1,0),(6,30,17,0),(6,30,21,0),
        (7,1,1,0),(7,1,16,0),(7,1,20,0),(7,2,0,0),(7,2,19,0),(7,2,23,0),
        (7,3,3,0),(7,3,18,0),(7,3,22,0),(7,4,1,30),(7,4,17,0),(7,4,21,0),
        (7,5,20,0),(7,6,0,0),(7,6,19,0),(7,7,0,0),(7,7,16,0),(7,7,20,0),
        (7,9,20,0),(7,10,19,0),(7,11,21,0),(7,12,1,0),
        (7,14,19,0),(7,15,19,0),(7,18,19,0),(7,19,19,0),
    ]:
        _live_sched.append([int(_cal.timegm((2026, _lmo2, _ld2, _lh2, _lmi2, 0))) * 1000,
                             "KO-Spiel", 210])
    live_sched_json = _json.dumps(_live_sched, ensure_ascii=False)

    # ── CSS (plain string – keine f-string-Escapes nötig) ─
    CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#152438;color:#eef5fd;min-height:100vh;font-size:15px}
.hdr{background:linear-gradient(135deg,#1e3f60,#0e2848);border-bottom:2px solid #c8a200;padding:16px 24px;display:flex;align-items:center;gap:14px}
.hdr-logo{height:60px;width:auto;filter:drop-shadow(0 2px 8px rgba(0,0,0,.6));flex-shrink:0}.hdr-title{font-size:1.5rem;font-weight:800;color:#f5c518}.hdr-sub{font-size:.82rem;color:#a0c0de;margin-top:3px}
.stats{display:flex;gap:12px;flex-wrap:wrap;padding:14px 24px;background:#182d45;border-bottom:1px solid #2e4e72}
.stat-box{background:#1c3450;border:1px solid #2e4e72;border-radius:8px;padding:10px 16px;min-width:130px}
.stat-val{font-size:1.4rem;font-weight:800;color:#f5c518}.stat-lbl{font-size:.75rem;color:#a0c0de;margin-top:2px}
.filter-bar{padding:12px 24px;background:#182d45;border-bottom:1px solid #2e4e72;display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.filter-lbl{color:#a0c0de;font-size:.82rem;font-weight:600;text-transform:uppercase;letter-spacing:.05em}
.drop-wrap{position:relative}
.drop-btn{background:#1c3450;border:1px solid #2e4e72;color:#eef5fd;padding:6px 14px;border-radius:6px;cursor:pointer;font-size:.82rem;min-width:210px;text-align:left;display:flex;align-items:center;justify-content:space-between;gap:10px;transition:all .2s}
.drop-btn:hover,.drop-btn.open{background:#2a4e78;border-color:#4fc3f7}
.drop-panel{display:none;position:absolute;top:calc(100% + 5px);left:0;min-width:230px;max-height:320px;background:#1c3450;border:1px solid #2e4e72;border-radius:8px;z-index:200;box-shadow:0 8px 28px rgba(0,0,0,.5);overflow:hidden}
.drop-panel.open{display:flex;flex-direction:column}
.drop-hdr{display:flex;border-bottom:1px solid #2e4e72;flex-shrink:0}
.drop-hdr-btn{flex:1;padding:7px;font-size:.75rem;color:#a0c0de;background:none;border:none;border-right:1px solid #2e4e72;cursor:pointer;transition:all .15s}
.drop-hdr-btn:last-child{border-right:none}.drop-hdr-btn:hover{color:#f5c518;background:#2a4e78}
.drop-opts{overflow-y:auto;flex:1}
.dlabel{display:flex;align-items:center;gap:9px;padding:8px 14px;cursor:pointer;font-size:.83rem;color:#eef5fd;transition:background .12s;user-select:none}
.dlabel:hover{background:#2a4e78}.dlabel input[type=checkbox]{accent-color:#4fc3f7;width:14px;height:14px;cursor:pointer;flex-shrink:0}
.hint{font-size:.75rem;color:#546e7a}
.section{padding:20px 24px}
.sec-title{font-size:.88rem;font-weight:700;color:#f5c518;text-transform:uppercase;letter-spacing:.08em;margin-bottom:14px;display:flex;align-items:center;gap:10px}
.sec-title::after{content:'';flex:1;height:1px;background:linear-gradient(to right,#1c3350,transparent)}
.wrap{overflow-x:auto}
.rtbl{width:100%;border-collapse:collapse;min-width:500px}
.rtbl th{background:#2a4e78;color:#f5c518;padding:9px 10px;text-align:right;font-size:.78rem;font-weight:600;white-space:nowrap}
.rtbl th:first-child,.rtbl th:nth-child(2){text-align:left}
.rtbl td{padding:8px 10px;text-align:right;border-bottom:1px solid #182d45;font-size:.88rem;transition:background .15s}
.rtbl td.pl,.rtbl td.nm{text-align:left}.rtbl td.pts{font-weight:700;font-size:1rem;color:#4fc3f7}
.rtbl tr{cursor:pointer}.rtbl tr:hover td{background:#1e3a5c!important}
.rtbl tr.sel td{box-shadow:inset 3px 0 0 #4fc3f7}
.rtbl tr.sel td.nm{color:#f5c518;font-weight:700}
.rtbl tr.r1 td{background:#3a2a00}.rtbl tr.r2 td{background:#1c2638}.rtbl tr.r3 td{background:#182a3e}
.no-sel{text-align:center;padding:44px 16px;color:#546e7a}
.no-sel-icon{font-size:2.4rem;margin-bottom:10px}.no-sel-txt{font-size:.9rem;color:#7a9bbe}
.badge{display:inline-block;padding:2px 9px;border-radius:12px;font-size:.72rem;font-weight:700;white-space:nowrap}
.badge.ex{background:#1b5e20;color:#a5d6a7;border:1px solid #2e7d32}
.badge.df{background:#004d40;color:#80cbc4;border:1px solid #00695c}
.badge.td{background:#bf360c;color:#ffccbc;border:1px solid #e64a19}
.badge.ms{background:#3e0000;color:#ef9a9a;border:1px solid #b71c1c}
.badge.op{background:#1c3048;color:#7a9bbe;border:1px solid #2e4a60}
.pcard{background:#182d45;border:1px solid #2e4e72;border-radius:10px;margin-bottom:18px;overflow:hidden}
.pcard.hidden{display:none}
.phdr{background:linear-gradient(90deg,#1e3f60,#182d45);padding:14px 18px;display:flex;align-items:center;gap:12px;cursor:pointer;user-select:none;border-bottom:1px solid #2e4e72}
.phdr:hover{background:#2a4e78}
.pmedal{font-size:1.6rem;flex-shrink:0}.pinfo{flex:1}.pname{font-size:1.05rem;font-weight:700;color:#f5c518}
.pbreakdown{display:flex;gap:8px;flex-wrap:wrap;margin-top:4px}
.pbitem{font-size:.72rem;color:#a0c0de}.pbval{color:#eef5fd;font-weight:600}
.ptotal{font-size:1.4rem;font-weight:800;color:#4fc3f7;flex-shrink:0}
.pbody{padding:16px 18px}
.gr-tabs{display:flex;flex-wrap:wrap;gap:5px;margin-bottom:12px}
.gtab{padding:3px 10px;background:#1c3450;border:1px solid #2e4e72;border-radius:5px;font-size:.75rem;cursor:pointer;color:#a0c0de;transition:all .15s}
.gtab:hover{background:#2a4e78}.gtab.active{background:#2a4e78;border-color:#4fc3f7;color:#f5c518;font-weight:700}
.gtbl{width:100%;border-collapse:collapse;min-width:450px}
.gtbl th{background:#1c3450;color:#7a9bbe;padding:7px 10px;font-size:.72rem;text-align:left;font-weight:600;text-transform:uppercase;white-space:nowrap}
.gtbl td{padding:7px 10px;border-bottom:1px solid #182d45;font-size:.84rem;vertical-align:middle}
.gtbl tr:hover td{background:#1e3a5c}.gtbl tr.gr-hide{display:none}
.sc{font-family:monospace;font-size:.9rem}.sc-t{color:#a0b8c8}.sc-r{color:#4fc3f7;font-weight:700}.arrow{color:#546e7a;margin:0 3px}
.ko-wrap{margin-top:16px;border-top:1px solid #2e4e72;padding-top:14px}
.ko-sec-title{font-size:.75rem;font-weight:700;color:#a0c0de;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px}
.ko-row{display:flex;align-items:flex-start;gap:10px;margin-bottom:8px}
.ko-rname{width:135px;font-size:.75rem;color:#7a9bbe;font-weight:600;text-transform:uppercase;padding-top:3px;flex-shrink:0}
.ko-rdate{font-size:.68rem;color:#546e7a;font-weight:400;text-transform:none}
.de-row td{background:rgba(245,197,24,.18)!important;border-left:4px solid #f5c518}
.de-row{box-shadow:inset 0 0 16px rgba(245,197,24,.10)}
.ko-teams{display:flex;flex-wrap:wrap;gap:5px}
.kt{padding:2px 10px;border-radius:12px;font-size:.75rem;background:#1c3450;border:1px solid #2e4e72;color:#eef5fd}
.kt.hit{background:#1b5e20;border-color:#2e7d32;color:#a5d6a7}.kt.miss{background:#3e0000;border-color:#b71c1c;color:#ef9a9a}
.foot{text-align:center;padding:18px;color:#546e7a;font-size:.78rem;border-top:1px solid #182d45}
.fb-btn{position:fixed;top:16px;right:16px;z-index:999;background:#f5c518;color:#0a1929;border:none;border-radius:20px;padding:7px 14px;font-size:.8rem;font-weight:700;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,.4);transition:background .2s}
.fb-btn:hover{background:#ffe066}
.fb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:1000;align-items:flex-start;justify-content:flex-end;padding:60px 16px 0}
.fb-overlay.open{display:flex}
.fb-box{background:#0e2035;border:1px solid #2e4e72;border-radius:12px;padding:18px 20px;width:320px;max-width:calc(100vw - 32px);box-shadow:0 8px 32px rgba(0,0,0,.6)}
.fb-box h3{margin:0 0 6px;font-size:1rem;color:#f5c518}
.fb-box p{margin:0 0 10px;font-size:.78rem;color:#7a9bbe}
.fb-box textarea{width:100%;box-sizing:border-box;background:#07151f;border:1px solid #2e4e72;border-radius:6px;color:#eef5fd;font-size:.85rem;padding:8px;resize:vertical;min-height:100px;outline:none}
.fb-box textarea:focus{border-color:#f5c518}
.fb-actions{display:flex;gap:8px;margin-top:10px}
.fb-send{flex:1;background:#f5c518;color:#0a1929;border:none;border-radius:8px;padding:8px;font-weight:700;cursor:pointer;font-size:.85rem}
.fb-send:hover{background:#ffe066}
.fb-cancel{background:#1c3450;color:#7a9bbe;border:1px solid #2e4e72;border-radius:8px;padding:8px 12px;cursor:pointer;font-size:.85rem}
.fb-cancel:hover{color:#eef5fd}
@media(max-width:600px){.section{padding:12px}.filter-bar{padding:10px 12px}.pbody{padding:12px}}
.gt-bar{display:flex;flex-wrap:wrap;gap:5px;padding:10px 0 8px}
.mtx-wrap{overflow-x:auto}
.mtx{width:100%;border-collapse:collapse;font-size:.82rem}
.mtx th{background:#1c3450;color:#a0c0de;padding:6px 8px;border:1px solid #2e4e72;text-align:center;white-space:nowrap;font-weight:600}
.mtx th.mh-l{text-align:left}
.mtx td{padding:5px 7px;border:1px solid #1e3550;vertical-align:middle;text-align:center}
.mtx td.mi{text-align:left;color:#7a9bbe;font-size:.78rem;white-space:nowrap;line-height:1.35}
.mi-time{font-size:.7rem;opacity:.65}
.gr-badge{display:inline-block;background:#1c3450;border:1px solid #2e4e72;color:#a0c0de;font-size:.65rem;font-weight:700;padding:1px 4px;border-radius:3px;margin-right:5px;vertical-align:middle}
.mtx td.mi-t{text-align:left;color:#c8d8e8;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:155px}
.mtx td.mi-r{text-align:center;font-weight:700;color:#f5c518;min-width:52px;max-width:52px}
.mtx tr.de-row td{background:rgba(245,197,24,.05)!important}
.mtx tr.kommend>td:first-child{border-left:3px solid #f5c518}
.mtx tr.kommend .mi{color:#f5c518}
.mcel{font-size:.82rem}
.mcel-ex{background:rgba(39,174,96,.22);color:#5dde8a}
.mcel-df{background:rgba(79,195,247,.15);color:#4fc3f7}
.mcel-td{background:rgba(255,167,38,.15);color:#ffa726}
.mcel-ms{background:rgba(180,40,40,.1);color:#7a9bbe}
.mcel-op{color:#c8d8e8}
.mcel-sub{display:block;font-size:.68rem;opacity:.8;margin-top:2px}
.mtx tfoot td{background:#182d45;border-top:2px solid #2e4e72;font-weight:700;color:#f5c518;text-align:center;padding:6px 8px}
.mtx tfoot td.sub-lbl{text-align:left;color:#7a9bbe;font-weight:400;font-size:.78rem}
.mpl-name{font-size:.8rem;font-weight:600}
.mpl-pts{font-size:.72rem;color:#f5c518;display:block;margin-top:2px}
.mtx-grp th:nth-child(-n+3){position:sticky;z-index:3;background:#1c3450}
.mtx-grp td:nth-child(-n+3){position:sticky;z-index:2;background:#152438}
.mtx-grp th:nth-child(1),.mtx-grp td:nth-child(1){left:0;min-width:60px;max-width:60px}
.mtx-grp th:nth-child(2),.mtx-grp td:nth-child(2){left:60px;min-width:155px;max-width:155px}
.mtx-grp th:nth-child(3),.mtx-grp td:nth-child(3){left:215px;min-width:52px;max-width:52px;box-shadow:4px 0 8px rgba(0,0,0,.45)}
.mtx-grp tr.de-row td:nth-child(-n+3){background:#1a2d42!important}
.mtx-grp tr.kommend td:nth-child(-n+3){background:#16202e!important}
.mtx-ko th:nth-child(-n+2){position:sticky;z-index:3;background:#1c3450}
.mtx-ko td:nth-child(-n+2){position:sticky;z-index:2;background:#152438}
.mtx-ko th:nth-child(1),.mtx-ko td:nth-child(1){left:0;min-width:36px;max-width:36px}
.mtx-ko th:nth-child(2),.mtx-ko td:nth-child(2){left:36px;box-shadow:4px 0 8px rgba(0,0,0,.45)}
.leg-box{background:#182d45;border:1px solid #2e4e72;border-radius:8px;padding:8px 14px;margin-bottom:10px;font-size:.78rem;cursor:pointer}
.leg-box summary{color:#a0c0de;font-weight:600;user-select:none;list-style:none}
.leg-box summary::-webkit-details-marker{display:none}
.leg-inner{margin-top:8px;display:flex;flex-direction:column;gap:6px;color:#7a9bbe;line-height:1.7}
.leg-inner span{color:#eef5fd}
.live-banner{display:none;background:linear-gradient(90deg,#1a0808,#250c0c);border-bottom:2px solid #c62828;padding:7px 24px;align-items:center;gap:10px;font-size:.83rem}.live-banner.visible{display:flex}
.live-dot{width:9px;height:9px;background:#e53935;border-radius:50%;flex-shrink:0;animation:pulse-live 1.2s ease-in-out infinite}
@keyframes pulse-live{0%,100%{opacity:1;transform:scale(1)}50%{opacity:.35;transform:scale(.65)}}
.live-txt{color:#ffcdd2;font-weight:700}.live-match{color:#ef9a9a}
.next-upd{font-size:.72rem;color:#546e7a;padding:3px 24px 5px;background:#0e1e30}
.chart-section{padding:16px 24px}.chart-wrap{background:#182d45;border:1px solid #2e4e72;border-radius:10px;padding:16px}
.chart-title{font-size:.85rem;font-weight:700;color:#f5c518;text-transform:uppercase;letter-spacing:.07em;margin-bottom:12px}
.arr-up{color:#66bb6a;font-size:.72rem;font-weight:700;margin-right:1px}
.arr-dn{color:#ef5350;font-size:.72rem;font-weight:700;margin-right:1px}
.arr-eq{color:#3a5f84;font-size:.72rem;margin-right:1px}
.print-btn{background:none;border:1px solid #2e4e72;border-radius:5px;color:#7a9bbe;padding:2px 7px;font-size:.72rem;cursor:pointer;margin-left:6px;transition:all .15s;vertical-align:middle}
.print-btn:hover{border-color:#4fc3f7;color:#4fc3f7}
#pvModal{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:2000;align-items:flex-start;justify-content:center;overflow-y:auto;padding:20px}
#pvModal.open{display:flex}
.pv-box{background:#fff;color:#111;border-radius:8px;padding:10mm;max-width:780px;width:100%;font-family:Arial,sans-serif;font-size:10pt;position:relative}
.pv-close{position:absolute;top:10px;right:14px;background:#e0e0e0;border:none;border-radius:5px;padding:4px 10px;cursor:pointer;font-size:.85rem;font-weight:700}
.pv-close:hover{background:#ccc}
.pv-print-action{text-align:right;margin-bottom:8px}
.pv-print-action button{background:#1a6ec7;color:#fff;border:none;border-radius:7px;padding:6px 14px;font-weight:700;cursor:pointer;font-size:.9rem}
.pv-print-action button:hover{background:#1d7fd6}
@media print{
  #pvModal{display:none!important}
  #pvModal.print-mode{display:block!important;position:static;background:none;padding:0}
  #pvModal.print-mode .pv-box{box-shadow:none;border-radius:0;padding:8mm 10mm;max-width:none}
  #pvModal.print-mode .pv-close,.pv-print-action{display:none!important}
  .hdr,.stats,.filter-bar,.section,.chart-section,.foot,.fb-btn,.next-upd,.live-banner{display:none!important}
}
.pvs-hdr{font-size:15pt;font-weight:800;margin-bottom:2pt;color:#1a3a5c}
.pvs-sub{font-size:8.5pt;color:#666;margin-bottom:12pt;border-bottom:1pt solid #ccc;padding-bottom:4pt}
.pvs-sec{font-size:11pt;font-weight:700;background:#dde8f4;padding:3pt 6pt;margin:8pt 0 3pt;color:#1a3a5c}
.pvs-grp{font-size:9pt;font-weight:600;background:#f0f4f8;padding:2pt 6pt;margin:4pt 0 1pt}
.pvs-row{font-size:8.5pt;display:flex;gap:8pt;padding:1.5pt 6pt;border-bottom:.3pt solid #e0e0e0}
.pvs-row:nth-child(odd){background:#fafafa}
.pvs-match{flex:1}.pvs-score{font-weight:700;min-width:35pt;color:#1a6ec7}
.pvs-date{color:#888;font-size:8pt;min-width:50pt}
.pvs-ko-hdr{font-size:8.5pt;font-weight:700;background:#edf2f8;padding:2pt 6pt;margin:5pt 0 1pt;color:#1a3a5c;border-left:3pt solid #1a6ec7}
.pvs-qs-grid{display:grid;grid-template-columns:1fr 1fr;gap:0;margin:0}
.pvs-qs-item{font-size:8pt;padding:1.5pt 6pt;border-bottom:.3pt solid #ebebeb;display:flex;gap:6pt}
.pvs-qs-src{color:#999;font-size:7.5pt;min-width:48pt;flex-shrink:0}
.pvs-ko-row{font-size:8.5pt;padding:2pt 6pt;display:flex;gap:8pt;border-bottom:.3pt solid #e0e0e0}
.pvs-ko-row:nth-child(odd){background:#fafafa}
.pvs-ko-lbl{width:105pt;color:#555;font-weight:600;flex-shrink:0}
.pvs-match-pair{font-size:8.5pt;padding:2pt 6pt;display:flex;gap:4pt;align-items:baseline;border-bottom:.3pt solid #e0e0e0}
.pvs-match-pair:nth-child(odd){background:#fafafa}
.pvs-mp-lbl{color:#999;font-size:7.5pt;min-width:32pt;flex-shrink:0}
.pvs-mp-teams{flex:1}
.pvs-mp-vs{color:#bbb;font-size:7.5pt;margin:0 3pt}
.pvs-mp-winner{min-width:70pt;font-weight:700;color:#1a5c1a;text-align:right;flex-shrink:0}
.pvs-leg{margin:0}.pvs-leg-row{font-size:8pt;padding:2pt 6pt;border-bottom:.3pt solid #ebebeb;line-height:1.5}
.pvs-leg-cat{font-weight:700;color:#1a3a5c;display:inline-block;min-width:105pt;margin-right:4pt}
@media(max-width:600px){
  .section,.chart-section{padding:12px}
  .filter-bar,.hdr{padding:10px 12px}
  .stats{padding:10px 12px;gap:8px}
  .stat-box{flex:1;min-width:calc(50% - 8px);padding:8px 10px}
  .stat-val{font-size:1.1rem}
  .rtbl th,.rtbl td{padding:5px 5px;font-size:.77rem}
  .mtx-grp th:nth-child(2),.mtx-grp td:nth-child(2){min-width:105px;max-width:105px}
  .next-upd{padding:3px 12px 5px}
  .fb-btn{top:10px;right:10px;padding:6px 11px;font-size:.75rem}}
"""

    # ── JavaScript (plain string – kein f-string-Escaping) ─
    JS = """
const MEDALS=["🥇","🥈","🥉"];
const GROUPS=["A","B","C","D","E","F","G","H","I","J","K","L"];
const KO_ROUNDS=[["GQ","Qualifikation fürs 16tel-Finale (32 Teams, je +3 Pkt)","28.06."],["S16","Achtelfinale-Teilnehmer (16 Teams, je +5 Pkt)","28.06.–04.07."],
  ["S8","Viertelfinale-Teilnehmer (8 Teams, je +10 Pkt)","04.–07.07."],
  ["VF","Halbfinale-Teilnehmer (4 Teams, je +15 Pkt)","09.–12.07."],
  ["P3","Platz 3 (Teilnehmer +10, Sieger +15)","18.07."],
  ["F","Finalisten (2 Teams, je +20 Pkt)","19.07."],["WM","Weltmeister (+25 Pkt)",""]];
// Bracket-Paarungen (spiegelt tippzettel.html, Quelle: FIFA FWC26 Regulations)
const S16_PAIRS=[[12,13],[2,17],[4,24],[5,14],[16,20],[8,25],[0,26],[11,27],[6,29],[3,28],[7,21],[22,23],[1,30],[15,18],[9,19],[10,31]];
const S8_PAIRS=[[2,5],[0,3],[1,4],[6,7],[11,10],[9,8],[14,13],[12,15]];
const VF_PAIRS=[[0,1],[4,5],[2,3],[6,7]];
let selected=new Set(DATA.players.map(p=>p.name));

function badge(pts){
  if(pts===null||pts===undefined) return '<span class="badge op">· Offen</span>';
  const cfg={4:["ex","⚽ Exakt +4"],3:["df","✓ Tordiff +3"],2:["td","≈ Tendenz +2"],0:["ms","✗ Daneben"]};
  const [cls,lbl]=cfg[pts]||["op","?"];
  return `<span class="badge ${cls}">${lbl}</span>`;
}

function renderDropdown(){
  const sorted=[...DATA.players].sort((a,b)=>a.name.localeCompare(b.name,"de"));
  const cnt=selected.size;
  const btn=document.getElementById("dropBtn");
  if(cnt===0)
    btn.innerHTML='Spieler auswählen <span style="opacity:.5;margin-left:4px">&#9662;</span>';
  else if(cnt===DATA.players.length)
    btn.innerHTML=`Alle (${cnt}) <span style="color:#4fc3f7;margin-left:4px">&#9662;</span>`;
  else
    btn.innerHTML=`<span style="color:#f5c518">${cnt} ausgewählt</span><span style="opacity:.5;margin-left:4px">&#9662;</span>`;
  document.getElementById("dropList").innerHTML=sorted.map(p=>
    `<label class="dlabel">
      <input type="checkbox" ${selected.has(p.name)?"checked":""}
        onchange="togglePlayerDrop('${p.name.replace(/'/g,"\\\\'")}',this.checked)">
      ${p.name}
    </label>`
  ).join("");
}

function toggleDropdown(e){
  e.stopPropagation();
  const panel=document.getElementById("dropPanel");
  const btn=document.getElementById("dropBtn");
  const open=panel.classList.toggle("open");
  btn.classList.toggle("open",open);
}

document.addEventListener("click",e=>{
  const wrap=document.getElementById("dropWrap");
  if(wrap&&!wrap.contains(e.target)){
    document.getElementById("dropPanel").classList.remove("open");
    document.getElementById("dropBtn").classList.remove("open");
  }
});

function togglePlayerDrop(name,checked){
  if(checked) selected.add(name); else selected.delete(name);
  renderAll();
}

function selectAll(){DATA.players.forEach(p=>selected.add(p.name));renderAll();}
function selectNone(){selected.clear();renderAll();}

function getRankArrow(name){
  if(!HISTORY||HISTORY.length<2)return '';
  const prev=HISTORY[HISTORY.length-2].pts,curr=HISTORY[HISTORY.length-1].pts;
  const rank=obj=>Object.entries(obj).sort((a,b)=>b[1]-a[1]).findIndex(([n])=>n===name);
  const pr=rank(prev),cr=rank(curr);
  if(pr<0||cr<0)return '';
  if(cr<pr)return '<span class="arr-up">↑</span>';
  if(cr>pr)return '<span class="arr-dn">↓</span>';
  return '<span class="arr-eq">=</span>';
}
function renderRanking(){
  document.getElementById("rankBody").innerHTML=DATA.players.map(p=>{
    const med=p.platz<=3?MEDALS[p.platz-1]:"";
    const cls=[p.platz<=3?`r${p.platz}`:"",selected.has(p.name)?"sel":""].filter(Boolean).join(" ");
    return `<tr class="${cls}" onclick="togglePlayerRow('${p.name.replace(/'/g,"\\\\'")}')">
      <td class="pl">${med}${getRankArrow(p.name)}${p.platz}.</td><td class="nm">${p.name} <button class="print-btn" onclick="event.stopPropagation();openPV('${p.name.replace(/'/g,"\\\\'")}')">🖨️</button></td>
      <td class="pts">${p.gesamt}</td><td>${p.gruppe}</td>
      <td>${p.gq}</td><td>${p.s16}</td><td>${p.s8}</td><td>${p.vf}</td>
      <td>${p.hf}</td><td>${p.p3}</td><td>${p.wm}</td>
    </tr>`;
  }).join("");
}

function togglePlayerRow(name){
  if(selected.has(name)) selected.delete(name); else selected.add(name);
  renderAll();
}

function findBestInitTab(){
  // Nach Gruppenphase: immer AF (16) als Default zeigen
  const filled=DATA.players[0]&&DATA.players[0].spiele.filter(s=>s.ergebnis).length;
  if(filled>=72) return 'S16';
  return 'Gruppen';
}
let activeTab=findBestInitTab();

function renderDetails(){
  const players=DATA.players.filter(p=>selected.has(p.name));
  const cont=document.getElementById("detailContainer");
  if(players.length===0){
    cont.innerHTML=`<div class="no-sel"><div class="no-sel-icon">👆</div>
      <div class="no-sel-txt">Spieler über das Dropdown auswählen</div></div>`;
    return;
  }
  const plHdrs=players.map(p=>{
    const med=p.platz<=3?MEDALS[p.platz-1]:`${p.platz}.`;
    return `<th class="mh-pl" style="min-width:75px"><span class="mpl-name">${med} ${p.name}</span><span class="mpl-pts">${p.gesamt} Pkt</span></th>`;
  }).join('');
  const tabsHtml=[
    `<span class="gtab${activeTab==='Gruppen'?' active':''}" onclick="setTab('Gruppen')">Gruppen</span>`,
    `<span style="color:#2e4e72;padding:0 4px">│</span>`,
    ...KO_ROUNDS.map(([k,lbl])=>{
      const short={GQ:'GQ (32)',S16:'AF (16)',S8:'VF (8)',VF:'HF (4)',P3:'Platz 3',F:'Finale',WM:'WM'}[k]||k;
      return `<span class="gtab${k===activeTab?' active':''}" onclick="setTab('${k}')">${short}</span>`;
    })
  ].join('');
  let tableHtml='';
  if(activeTab==='Gruppen'){
    const sortKey=s=>{
      const p=s.datum.replace(/\\.$/,'').split('.');
      const[h,mn]=(s.uhrzeit||'00:00').split(':').map(Number);
      return parseInt(p[1])*1e6+parseInt(p[0])*1e4+h*100+mn;
    };
    const allMatches=[...players[0].spiele].sort((a,b)=>sortKey(a)-sortKey(b));
    const rows=allMatches.map(s=>{
      const hasRes=s.ergebnis&&s.ergebnis!=='';
      const isDE=s.heim==='Deutschland'||s.gast==='Deutschland';
      const trCls=[isDE?'de-row':'',s.kommend?'kommend':''].filter(Boolean).join(' ');
      const plCells=players.map(p=>{
        const ps=p.spiele.find(x=>x.id===s.id)||{};
        const tipp=ps.tipp||'–';
        const pts=hasRes?ps.punkte:null;
        // Keine vorläufigen Punkte während laufender Spiele
        const cls=pts===null?'mcel-op':pts===4?'mcel-ex':pts===3?'mcel-df':pts===2?'mcel-td':'mcel-ms';
        const sub=pts!==null?`<span class="mcel-sub">${pts>0?'+'+pts:'0'}</span>`:'';
        return `<td class="mcel ${cls}">${tipp}${sub}</td>`;
      }).join('');
      const datumHtml=`${s.datum}${s.uhrzeit?`<br><span class="mi-time">${s.uhrzeit}</span>`:''}`;
      const spielHtml=`<span class="gr-badge">${s.gr}</span>${s.heim} <span style="color:#3a5f84">vs</span> ${s.gast}`;
      const hasLive=s.live&&s.live!=='';
      const ergHtml=hasRes?s.ergebnis:hasLive?`<span style="color:#e53935;font-weight:700">▶ ${s.live}</span>`:s.kommend?'▶':'–';
      return `<tr${trCls?` class="${trCls}"`:''}><td class="mi">${datumHtml}</td><td class="mi-t">${spielHtml}</td><td class="mi-r">${ergHtml}</td>${plCells}</tr>`;
    }).join('');
    const subs=players.map(p=>{
      const pts=p.spiele.filter(s=>s.punkte!==null&&s.punkte!==undefined).reduce((a,s)=>a+(s.punkte||0),0);
      const n=p.spiele.filter(s=>s.ergebnis).length;
      return `<td><b>${pts}</b><span style="opacity:.5;font-size:.73rem;margin-left:4px">(${n}/72)</span></td>`;
    }).join('');
    tableHtml=`<div class="mtx-wrap"><table class="mtx mtx-grp"><thead><tr><th class="mh-l">Datum</th><th class="mh-l">Spiel</th><th>Erg.</th>${plHdrs}</tr></thead><tbody>${rows}</tbody><tfoot><tr><td colspan="3" class="sub-lbl">Gruppenphase gesamt:</td>${subs}</tr></tfoot></table></div>`;
  } else if(activeTab==='GQ'){
    const gqActual=DATA.ko['GQ']||[];
    const gqActualLow=gqActual.map(t=>t.toLowerCase());
    const hasActual=gqActual.length>0;
    const allPredTeams=new Set();
    players.forEach(p=>(p.gq_teams||[]).forEach(t=>{if(t)allPredTeams.add(t)}));
    const teamList=hasActual ? gqActual : [...allPredTeams].sort((a,b)=>a.localeCompare(b,'de'));
    const rows=teamList.length===0
      ?`<tr><td colspan="${players.length+2}" style="text-align:center;padding:24px;color:#546e7a">Keine Daten</td></tr>`
      :teamList.map((team,i)=>{
        const teamLow=team.toLowerCase();
        const plCells=players.map(p=>{
          const pred=(p.gq_teams||[]).map(t=>t.toLowerCase());
          const tipped=pred.includes(teamLow);
          if(!hasActual){
            return `<td class="mcel ${tipped?'mcel-ex':'mcel-op'}">${tipped?'✓':'–'}</td>`;
          }
          const hit=tipped;
          return `<td class="mcel ${hit?'mcel-ex':'mcel-ms'}">${hit?'✓':'✗'}</td>`;
        }).join('');
        return `<tr><td class="mi" style="white-space:nowrap;padding:5px 8px;min-width:30px;font-size:.78rem;color:#7a9bbe">${i+1}.</td><td class="mi-t">${team}</td>${plCells}</tr>`;
      }).join('');
    const subs=players.map(p=>{
      const predLow=(p.gq_teams||[]).map(t=>t.toLowerCase());
      const hits=predLow.filter(t=>gqActualLow.includes(t)).length;
      const tipCount=(p.gq_teams||[]).length;
      return hasActual
        ?`<td><b>${hits*3}</b><span style="opacity:.5;font-size:.73rem;margin-left:4px">(${hits}/32)</span></td>`
        :`<td><span style="opacity:.5;font-size:.73rem">${tipCount}/32 getippt</span></td>`;
    }).join('');
    tableHtml=`<div class="mtx-wrap"><table class="mtx mtx-ko"><thead><tr><th style="min-width:30px;text-align:left;padding-left:8px"></th><th class="mh-l">Qualifikant fürs 16tel-Finale (je +3 Pkt)</th>${plHdrs}</tr></thead><tbody>${rows}</tbody><tfoot><tr><td></td><td class="sub-lbl">GQ gesamt:</td>${subs}</tr></tfoot></table></div>`;
  } else {
    const actual=DATA.ko[activeTab]||[];
    const norm=actual.map(t=>(t||'').toLowerCase());
    if(activeTab==='WM'){
      const plCells=players.map(p=>{
        const tip=p.ko_tipps['WM']||'–';
        const hit=norm.length>0&&tip.toLowerCase()===norm[0];
        const miss=norm.length>0&&!hit;
        return `<td class="mcel ${hit?'mcel-ex':miss?'mcel-ms':'mcel-op'}">${tip}</td>`;
      }).join('');
      tableHtml=`<div class="mtx-wrap"><table class="mtx mtx-ko"><thead><tr><th class="mh-l">Weltmeister</th>${plHdrs}</tr></thead><tbody><tr><td class="mi-t">${actual[0]||'–'}</td>${plCells}</tr></tbody></table></div>`;
    } else if(activeTab==='P3'){
      const p3parts=(DATA.ko_display&&DATA.ko_display['P3_participants'])||[];
      const p3t1=p3parts[0]||'–';
      const p3t2=p3parts[1]||'–';
      const winnerActual=(DATA.ko_display&&DATA.ko_display['P3']&&DATA.ko_display['P3'][0])||'–';
      const winnerNorm=winnerActual!=='–'?winnerActual.toLowerCase():'';
      const p3t1Norm=p3t1!=='–'?p3t1.toLowerCase():'';
      const p3t2Norm=p3t2!=='–'?p3t2.toLowerCase():'';
      // Getippte P3-Teilnehmer pro Spieler = VF-Tipps minus HF-Tipps
      function getPlayerP3Parts(p){
        const vf=(p.ko_tipps['VF']||[]).filter(Boolean);
        const hfNorm=(p.ko_tipps['HF']||[]).map(t=>(t||'').toLowerCase()).filter(Boolean);
        return vf.filter(t=>!hfNorm.includes((t||'').toLowerCase()));
      }
      const partCells1=players.map(p=>{
        const parts=getPlayerP3Parts(p);
        const tip=parts[0]||'–';
        const hit=p3t1Norm&&tip.toLowerCase()===p3t1Norm;
        const miss=p3t1Norm&&!hit&&tip!=='–';
        const cls=hit?'mcel-ex':miss?'mcel-ms':'mcel-op';
        return `<td class="mcel ${cls}">${tip}</td>`;
      }).join('');
      const partCells2=players.map(p=>{
        const parts=getPlayerP3Parts(p);
        const tip=parts[1]||'–';
        const hit=p3t2Norm&&tip.toLowerCase()===p3t2Norm;
        const miss=p3t2Norm&&!hit&&tip!=='–';
        const cls=hit?'mcel-ex':miss?'mcel-ms':'mcel-op';
        return `<td class="mcel ${cls}">${tip}</td>`;
      }).join('');
      const rowP3part1=`<tr><td class="mi" style="color:#7a9bbe;font-size:.78rem;padding:5px 8px">Teilnehmer 1</td><td class="mi-t">${p3t1}</td>${partCells1}</tr>`;
      const rowP3part2=`<tr><td class="mi" style="color:#7a9bbe;font-size:.78rem;padding:5px 8px">Teilnehmer 2</td><td class="mi-t">${p3t2}</td>${partCells2}</tr>`;
      const siegerCells=players.map(p=>{
        const tip=p.ko_tipps['P3']||'–';
        const tipLow=tip.toLowerCase();
        const isWinner=winnerNorm&&tipLow===winnerNorm;
        const isParticipant=(p3t1Norm||p3t2Norm)&&(tipLow===p3t1Norm||tipLow===p3t2Norm);
        const known=winnerNorm||(p3t1Norm||p3t2Norm);
        const cls=isWinner?'mcel-ex':isParticipant?'mcel-td':known?'mcel-ms':'mcel-op';
        return `<td class="mcel ${cls}">${tip}</td>`;
      }).join('');
      const rowSieger=`<tr><td class="mi" style="color:#7a9bbe;font-size:.78rem;padding:5px 8px">Sieger-Tipp</td><td class="mi-t">${winnerActual}</td>${siegerCells}</tr>`;
      tableHtml=`<div class="mtx-wrap"><table class="mtx mtx-ko"><thead><tr><th style="min-width:80px;text-align:left;padding-left:8px"></th><th class="mh-l">Platz 3 (Teilnehmer +10, Sieger +15)</th>${plHdrs}</tr></thead><tbody>${rowP3part1}${rowP3part2}${rowSieger}</tbody></table></div>`;
    } else {
      const tips=players.map(p=>Array.isArray(p.ko_tipps[activeTab])?p.ko_tipps[activeTab]:[]);
      const maxLen=Math.max(...tips.map(t=>t.length),0);
      const isFinale=activeTab==='F';
      const rowLabels=isFinale?['Finalist 1','Finalist 2']:null;
      const colHdrMap={S16:'Achtelfinale-Teilnehmer (je +5 Pkt)',S8:'Viertelfinale-Teilnehmer (je +10 Pkt)',VF:'Halbfinale-Teilnehmer (je +15 Pkt)',F:'Finalist (je +20 Pkt)'};
      const colHdr=colHdrMap[activeTab]||'Weitergekommen';
      const rows=maxLen===0
        ?`<tr><td colspan="${players.length+2}" style="text-align:center;padding:24px;color:#546e7a">Runde noch nicht begonnen</td></tr>`
        :Array.from({length:maxLen},(_,i)=>{
          const plCells=players.map((p,pi)=>{
            const tip=tips[pi][i]||'–';
            const hit=norm.includes(tip.toLowerCase())&&norm.length>0;
            const miss=norm.length>0&&!hit&&tip!=='–';
            return `<td class="mcel ${hit?'mcel-ex':miss?'mcel-ms':'mcel-op'}">${tip}</td>`;
          }).join('');
          const rowLbl=rowLabels?rowLabels[i]:`${i+1}.`;
          return `<tr><td class="mi" style="white-space:nowrap;padding:5px 8px;min-width:80px;font-size:.78rem;color:#7a9bbe">${rowLbl}</td><td class="mi-t">${actual[i]||'–'}</td>${plCells}</tr>`;
        }).join('');
      const ptsPerHit={'S16':5,'S8':10,'VF':15,'HF':20,'F':20}[activeTab]||5;
      const subs=norm.length>0?players.map(p=>{
        const tipsLow=(Array.isArray(p.ko_tipps[activeTab])?p.ko_tipps[activeTab]:[]).map(t=>(t||'').toLowerCase());
        const hits=tipsLow.filter(t=>t&&norm.includes(t)).length;
        return `<td><b>${hits*ptsPerHit}</b><span style="opacity:.5;font-size:.73rem;margin-left:4px">(${hits}/${norm.length})</span></td>`;
      }).join(''):'';
      const tfoot=norm.length>0?`<tfoot><tr><td colspan="2" class="sub-lbl">${colHdr.split('(')[0].trim()} gesamt:</td>${subs}</tr></tfoot>`:'';
      tableHtml=`<div class="mtx-wrap"><table class="mtx mtx-ko"><thead><tr><th style="min-width:80px;text-align:left;padding-left:8px"></th><th class="mh-l">${colHdr}</th>${plHdrs}</tr></thead><tbody>${rows}</tbody>${tfoot}</table></div>`;
    }
  }
  cont.innerHTML=`<div class="gt-bar">${tabsHtml}</div><details class="leg-box"><summary>📋 Punktesystem</summary><div class="leg-inner"><div><span style="color:#a0c0de;font-weight:600">Gruppenphase: </span><span class="badge ex">⚽ Exakt +4</span> <span class="badge df">✓ Tordiff +3</span> <span class="badge td">≈ Tendenz +2</span> <span class="badge ms">✗ Daneben 0</span></div><div><span style="color:#a0c0de;font-weight:600">KO-Runden: </span>Gr.-Qual. (32 Teams) +3 · Achtelfinale (16 Teams) +5 · Viertelfinale (8 Teams) +10 · Halbfinale (4 Teams) +15 · Platz 3 (im Spiel +10, Sieger +15) · Finale +20/Team · Weltmeister +25</div></div></details>${tableHtml}`;
}

function setTab(tab){activeTab=tab;renderDetails();}

function buildPlayerPV(p){
  const GROUPS_PV=["A","B","C","D","E","F","G","H","I","J","K","L"];
  let html=`<div class="pvs-hdr">WM 2026 Tippspiel – ${p.name}</div>`;
  html+=`<div class="pvs-sub">Platz ${p.platz} · ${p.gesamt} Pkt · Tippzettel</div>`;

  // Gruppenphase
  html+='<div class="pvs-sec">Gruppenphase &nbsp;·&nbsp; Exakt +4 | Tordiff +3 | Tendenz +2 | Daneben 0</div>';
  const byGr={};
  for(const s of p.spiele){(byGr[s.gr]=byGr[s.gr]||[]).push(s);}
  for(const g of GROUPS_PV){
    if(!byGr[g])continue;
    html+=`<div class="pvs-grp">Gruppe ${g}</div>`;
    for(const s of byGr[g]){
      html+=`<div class="pvs-row"><span class="pvs-date">${s.datum} ${s.uhrzeit||''}</span><span class="pvs-match">${s.heim} vs ${s.gast}</span><span class="pvs-score">${s.tipp||'–'}</span></div>`;
    }
  }

  // GQ
  html+='<div class="pvs-sec">Gruppenqualifikanten &nbsp;·&nbsp; 32 Teams · je +3 Pkt</div>';
  html+='<div class="pvs-qs-grid">';
  const gqActual=DATA.ko['GQ']||[];
  (p.gq_teams||[]).forEach(t=>{
    const hit=gqActual.includes(t.toLowerCase());
    html+=`<div class="pvs-qs-item" style="${hit?'color:#1b5e20;font-weight:700':'color:#888'}">${t}</div>`;
  });
  html+='</div>';

  // KO-Runden
  html+='<div class="pvs-sec">KO-Runden</div>';

  function pvMatchRow(label,tA,tB,winner,actualA,actualB){
    const norm=s=>s?s.toLowerCase().trim():'';
    const wA=winner&&norm(winner)===norm(tA);
    const wB=winner&&norm(winner)===norm(tB);
    const hitA=actualA&&actualA.includes(norm(tA));
    const hitB=actualB&&actualB.includes(norm(tB));
    const styleA=wA?'font-weight:800;color:#1a5c1a':(tA&&tA!=='?'?'':(tA=tA||'?','color:#bbb'));
    const styleB=wB?'font-weight:800;color:#1a5c1a':(tB&&tB!=='?'?'':(tB=tB||'?','color:#bbb'));
    const check=winner?' <span style="color:#2e7d32">✓</span>':'';
    return `<div class="pvs-match-pair"><span class="pvs-mp-lbl">${label}</span>`
      +`<span class="pvs-mp-teams"><span style="${styleA}">${tA||'?'}</span>`
      +` <span class="pvs-mp-vs">vs</span> `
      +`<span style="${styleB}">${tB||'?'}</span></span>`
      +`<span class="pvs-mp-winner">${winner?winner+check:'–'}</span></div>`;
  }

  const koTipps=p.ko_tipps;
  const s16tips=Array.isArray(koTipps['S16'])?koTipps['S16']:[];
  const s8tips =Array.isArray(koTipps['S8']) ?koTipps['S8'] :[];
  const vftips =Array.isArray(koTipps['VF']) ?koTipps['VF'] :[];
  const hftips =Array.isArray(koTipps['HF']) ?koTipps['HF'] :[];
  const actualS16=(DATA.ko['S16']||[]);
  const actualS8 =(DATA.ko['S8'] ||[]);
  const actualVF =(DATA.ko['VF'] ||[]);
  const actualHF =(DATA.ko['HF'] ||[]);

  // S16 – GQ-Teams als Basis (Qualifizierer-Reihenfolge), Sieger = s16tips
  const gqAll=p.gq_teams||[];
  const cnt16=s16tips.filter(t=>t&&t!=='').length;
  html+=`<div class="pvs-ko-hdr">Sechzehntelfinale · 16 Spiele · Weiterkommer je +5 Pkt · ${cnt16}/16 getippt</div>`;
  S16_PAIRS.forEach(([ai,bi],i)=>{
    const tA=gqAll[ai]||'?', tB=gqAll[bi]||'?';
    html+=pvMatchRow(`Sp.${i+1}`,tA,tB,s16tips[i]||null,actualS16,actualS16);
  });

  // S8 – s16tips als Basis
  const cnt8=s8tips.filter(t=>t&&t!=='').length;
  html+=`<div class="pvs-ko-hdr">Achtelfinale · 8 Spiele · Weiterkommer je +10 Pkt · ${cnt8}/8 getippt</div>`;
  S8_PAIRS.forEach(([ai,bi],i)=>{
    const tA=s16tips[ai]||'?', tB=s16tips[bi]||'?';
    html+=pvMatchRow(`Sp.${i+1}`,tA,tB,s8tips[i]||null,actualS8,actualS8);
  });

  // VF – s8tips als Basis
  const cntVF=vftips.filter(t=>t&&t!=='').length;
  html+=`<div class="pvs-ko-hdr">Viertelfinale · 4 Spiele · Weiterkommer je +15 Pkt · ${cntVF}/4 getippt</div>`;
  VF_PAIRS.forEach(([ai,bi],i)=>{
    const tA=s8tips[ai]||'?', tB=s8tips[bi]||'?';
    html+=pvMatchRow(`Sp.${i+1}`,tA,tB,vftips[i]||null,actualVF,actualVF);
  });

  // HF – vftips[0,1] vs vftips[2,3]
  const cntHF=hftips.filter(t=>t&&t!=='').length;
  html+=`<div class="pvs-ko-hdr">Halbfinale · 2 Spiele · Weiterkommer je +20 Pkt · ${cntHF}/2 getippt</div>`;
  [[0,1],[2,3]].forEach(([ai,bi],i)=>{
    const tA=vftips[ai]||'?', tB=vftips[bi]||'?';
    html+=pvMatchRow(`Sp.${i+1}`,tA,tB,hftips[i]||null,actualHF,actualHF);
  });

  // Platz 3
  const p3participants=(DATA.ko_display&&DATA.ko_display['P3_participants'])||[];
  const p3winner=(DATA.ko_display&&DATA.ko_display['P3']&&DATA.ko_display['P3'][0])||'';
  const p3tip=koTipps['P3']||null;
  html+='<div class="pvs-ko-hdr">Spiel um Platz 3 · Teilnahme +10 Pkt · Sieger +15 Pkt</div>';
  html+=pvMatchRow('Platz 3',p3participants[0]||'–',p3participants[1]||'–',p3tip,[p3winner.toLowerCase()],[p3winner.toLowerCase()]);

  // Finale
  const ftip=Array.isArray(koTipps['HF'])?null:null;
  const wmtip=koTipps['WM']||null;
  const actualF=(DATA.ko['HF']||[]);
  html+='<div class="pvs-ko-hdr">Finale · Finalist richtig je +20 Pkt · Weltmeister +25 Pkt</div>';
  html+=pvMatchRow('Finale',hftips[0]||'?',hftips[1]||'?',wmtip,(DATA.ko['F']||[]),(DATA.ko['F']||[]));

  html+='<div class="pvs-sec">Punktesystem</div>';
  html+=`<div class="pvs-leg">
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Gruppenphase:</span>Exakt +4 · Tordifferenz +3 · Tendenz richtig +2 · Daneben 0 Pkt</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Gr.-Qualifikanten:</span>32 Teams, je +3 Pkt pro richtig getipptem Qualifikanten</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Sechzehntelfinale:</span>16 Weiterkommer, je +5 Pkt (Sieger-Tipp grün hervorgehoben)</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Achtelfinale:</span>8 Weiterkommer, je +10 Pkt</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Viertelfinale:</span>4 Weiterkommer, je +15 Pkt</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Halbfinale:</span>2 Weiterkommer, je +20 Pkt</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Platz 3:</span>Tipp auf Teilnehmer +10 Pkt · Sieger richtig +15 Pkt (max. 25 Pkt)</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Finalisten:</span>Beide richtig +40 · Ein Finalist richtig +20 · Kein Treffer 0 Pkt</div>
    <div class="pvs-leg-row"><span class="pvs-leg-cat">Weltmeister:</span>+25 Pkt</div>
  </div>`;

  return html;
}

function openPV(name){
  const p=DATA.players.find(x=>x.name===name);
  if(!p)return;
  const modal=document.getElementById('pvModal');
  document.getElementById('pvContent').innerHTML=buildPlayerPV(p);
  modal.classList.add('open');
}
function closePV(){
  const modal=document.getElementById('pvModal');
  modal.classList.remove('open','print-mode');
}
function printPV(){
  const modal=document.getElementById('pvModal');
  modal.classList.add('print-mode');
  window.print();
  setTimeout(()=>modal.classList.remove('print-mode'),500);
}

function checkLiveMatch(){
  const now=Date.now();
  const active=LIVE_SCHED.find(s=>now>=s[0]&&now<=s[0]+s[2]*60000);
  const banner=document.getElementById('liveBanner');
  const matchTxt=document.getElementById('liveMatch');
  if(active){
    banner.classList.add('visible');
    // Live-Score aus DATA.live anzeigen falls vorhanden
    const liveScores=DATA.live||{};
    const liveEntry=Object.entries(liveScores).find(([id,v])=>{
      const sp=DATA.players[0]&&DATA.players[0].spiele.find(s=>s.id===id);
      return sp&&(active[1].includes(sp.heim)||active[1].includes(sp.gast));
    });
    if(liveEntry){
      const sp=DATA.players[0].spiele.find(s=>s.id===liveEntry[0]);
      // liveEntry[1] = "0:0 (45'+4')" → Score und Minute trennen
      const liveRaw=liveEntry[1];
      const minuteMatch=liveRaw.match(/\(([^)]+)\)/);
      const score=liveRaw.replace(/\s*\([^)]+\)/,'').trim();
      const minute=minuteMatch?` · ${minuteMatch[1]}`:'';
      matchTxt.textContent=`${sp.heim} ${score} ${sp.gast}${minute}`;
    } else {
      matchTxt.textContent=active[1];
    }
  }
  else{banner.classList.remove('visible');}
}
function renderNextUpdate(){
  const el=document.getElementById('nextUpd');
  if(!el)return;
  const now=Date.now();
  const future=LIVE_SCHED.filter(s=>s[0]>now).sort((a,b)=>a[0]-b[0]);
  if(!future.length){el.textContent='';return;}
  const next=future[0];
  const d=new Date(next[0]);
  const dateStr=d.toLocaleString('de-DE',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit',timeZone:'Europe/Berlin'});
  el.textContent='Nächstes Spiel: '+next[1]+' · '+dateStr+' Uhr (MESZ) · Update ~30 Min nach Spielende';
}
function renderChart(){
  if(!HISTORY||HISTORY.length<2)return;
  const el=document.getElementById('chartWrap');
  if(!el)return;
  // Nur die letzten 10 Tage anzeigen – eindeutige Tage extrahieren
  const allDays=[...new Set(HISTORY.map(h=>h.ts.split(' ')[0]))];
  const last10Days=allDays.slice(-10);
  // Pro Tag nur den letzten Eintrag behalten
  const filtered=last10Days.map(day=>{
    const entries=HISTORY.filter(h=>h.ts.startsWith(day));
    return entries[entries.length-1];
  }).filter(Boolean);
  // Nur einmalige Punkte-Änderungen zeigen (dünne History trotzdem vollständig)
  const labels=filtered.map(h=>h.ts);
  const allNames=Object.keys(HISTORY[HISTORY.length-1].pts||{});
  // Rangfolge nach aktuellem Stand für Farbreihenfolge
  const sortedNames=[...allNames].sort((a,b)=>(HISTORY[HISTORY.length-1].pts[b]||0)-(HISTORY[HISTORY.length-1].pts[a]||0));
  const colors=['#f5c518','#c0c0c0','#cd7f32','#4fc3f7','#66bb6a','#ef5350','#ab47bc','#ffa726','#26c6da','#d4e157','#ec407a','#ff7043'];
  const datasets=sortedNames.map((name,i)=>({
    label:name,
    data:filtered.map(h=>h.pts[name]!==undefined?h.pts[name]:null),
    borderColor:colors[i%colors.length],
    backgroundColor:colors[i%colors.length]+'22',
    borderWidth:i<3?2.5:1.5,
    tension:.35,
    pointRadius:3,
    spanGaps:true
  }));
  el.innerHTML='<canvas id="ptChart"></canvas>';
  new Chart(document.getElementById('ptChart'),{
    type:'line',
    data:{labels,datasets},
    options:{
      responsive:true,maintainAspectRatio:false,
      plugins:{
        legend:{
          labels:{color:'#a0c0de',font:{size:11}},
          onClick:(e,item,legend)=>{
            const ci=legend.chart;
            const meta=ci.getDatasetMeta(item.datasetIndex);
            meta.hidden=!meta.hidden;
            ci.update();
          }
        }
      },
      scales:{
        x:{ticks:{color:'#546e7a',font:{size:10},maxRotation:30},grid:{color:'#1e3550'}},
        y:{ticks:{color:'#546e7a',font:{size:10}},grid:{color:'#1e3550'},beginAtZero:false}
      }
    }
  });
}
function renderAll(){renderRanking();renderDropdown();renderDetails();}
renderAll();
checkLiveMatch();
renderNextUpdate();
renderChart();
setInterval(checkLiveMatch,60000);
"""

    # ── HTML-Gerüst (f-string nur für dynamische Werte) ───
    html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>WM 2026 Tippspiel</title>
<style>{CSS}</style>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
</head>
<body>
<div class="hdr">
  <img class="hdr-logo" src="https://upload.wikimedia.org/wikipedia/en/thumb/1/17/2026_FIFA_World_Cup_emblem.svg/120px-2026_FIFA_World_Cup_emblem.svg.png" alt="WM 2026 Logo">
  <div><div class="hdr-title">WM 2026 Tippspiel</div>
  <div class="hdr-sub">Stand: {ts} · {filled}/72 Gruppenspiele</div></div>
</div>
</div>
<div class="live-banner" id="liveBanner"><span class="live-dot"></span><span class="live-txt">LIVE:</span>&nbsp;<span class="live-match" id="liveMatch"></span></div>
<div class="stats">
  <div class="stat-box"><div class="stat-val">{n}</div><div class="stat-lbl">Teilnehmer</div></div>
  <div class="stat-box"><div class="stat-val">{leader}</div><div class="stat-lbl">🏆 Führend</div></div>
  <div class="stat-box"><div class="stat-val">{lpts} Pkt</div><div class="stat-lbl">Höchstpunktzahl</div></div>
  <div class="stat-box"><div class="stat-val">{filled}/72</div><div class="stat-lbl">Spiele ausgewertet</div></div>
</div>
<div id="nextUpd" class="next-upd"></div>
<div class="filter-bar">
  <span class="filter-lbl">Detailansicht:</span>
  <div class="drop-wrap" id="dropWrap">
    <button class="drop-btn" id="dropBtn" onclick="toggleDropdown(event)">Spieler auswählen &#9662;</button>
    <div class="drop-panel" id="dropPanel">
      <div class="drop-hdr">
        <button class="drop-hdr-btn" onclick="selectAll()">Alle ausw.</button>
        <button class="drop-hdr-btn" onclick="selectNone()">Keine</button>
      </div>
      <div class="drop-opts" id="dropList"></div>
    </div>
  </div>
  <span class="hint">&#183; Zeile anklicken f&#252;r Details / Vergleich</span>
</div>
<div class="section">
  <div class="sec-title">Gesamtrangliste</div>
  <div class="wrap">
    <table class="rtbl">
      <thead><tr>
        <th>Platz</th><th>Name</th><th>Gesamt</th><th>Gruppe</th>
        <th>GQ</th><th>AF</th><th>VF</th><th>HF</th><th>P3</th><th>WM</th>
      </tr></thead>
      <tbody id="rankBody"></tbody>
    </table>
  </div>
</div>
<div class="chart-section"><div class="chart-wrap"><div class="chart-title">Punkteverlauf</div><div id="chartWrap" style="height:260px;position:relative"></div></div></div>
<div class="section">
  <div class="sec-title">Detailauswertung</div>
  <div id="detailContainer"></div>
</div>
<div class="foot">Generiert am {ts} · WM 2026 Tippspiel</div>
<button class="fb-btn" onclick="document.getElementById('fbOverlay').classList.add('open')">💬 Feedback</button>
<div id="pvModal" onclick="if(event.target===this)closePV()">
  <div class="pv-box">
    <div class="pv-print-action"><button onclick="printPV()">🖨️ Als PDF speichern</button></div>
    <button class="pv-close" onclick="closePV()">✕ Schließen</button>
    <div id="pvContent"></div>
  </div>
</div>
<div class="fb-overlay" id="fbOverlay" onclick="if(event.target===this)closeFb()">
  <div class="fb-box">
    <h3>Feedback &amp; Verbesserungsvorschläge</h3>
    <p>Ideen, Wünsche oder Fehler? Schreib es kurz auf – öffnet deinen Mail-Client.</p>
    <textarea id="fbText" placeholder="Dein Feedback..."></textarea>
    <div class="fb-actions">
      <button class="fb-send" onclick="sendFb()">📧 Senden</button>
      <button class="fb-cancel" onclick="closeFb()">Abbrechen</button>
    </div>
  </div>
</div>
<script>
const DATA={data_json};
const HISTORY={history_json};
const LIVE_SCHED={live_sched_json};
{JS}
function closeFb(){{document.getElementById('fbOverlay').classList.remove('open');}}
function sendFb(){{
  const txt=document.getElementById('fbText').value.trim();
  if(!txt)return;
  window.location.href='mailto:timo.mayer@sap.com?subject=WM%202026%20Tippspiel%20-%20Feedback&body='+encodeURIComponent(txt);
  closeFb();
}}
</script>
</body></html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

# ============================================================
# MAIN
# ============================================================

def main():
    if len(sys.argv) < 3:
        print("Verwendung: python auswertung.py <tipps_ordner_oder_datei> <ergebnisse.xlsx>")
        print("Beispiel:   python auswertung.py Tipps/  Ergebnisse.xlsx")
        sys.exit(1)

    tipps_path      = Path(sys.argv[1])
    ergebnisse_path = Path(sys.argv[2])

    print(f"Lese Tipps aus:      {tipps_path}")
    print(f"Lese Ergebnisse aus: {ergebnisse_path}")

    if tipps_path.is_dir() or str(tipps_path).lower().endswith(".json"):
        df = read_json_tipps(tipps_path)
    else:
        df = read_responses(tipps_path)

    gs_results, ko_results = read_ergebnisse(ergebnisse_path)

    # Live-Scores einlesen (falls vorhanden)
    live_path = Path(__file__).parent / "live_scores.json"
    gs_live = {}
    if live_path.exists():
        try:
            gs_live = json.loads(live_path.read_text(encoding="utf-8"))
        except Exception:
            gs_live = {}

    print(f"Teilnehmer gefunden: {len(df)}")
    print(f"Gruppenspiel-Ergebnisse eingetragen: {sum(1 for v in gs_results.values() if v)}/72")
    if gs_live:
        print(f"Live-Spiele: {len(gs_live)}")

    rows = auswerten(df, gs_results, ko_results, gs_live=gs_live)

    out_path = Path("Rangliste.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_rangliste(wb, rows)
    write_details(wb, rows, gs_results)
    write_ko_details(wb, rows, ko_results)

    wb.save(out_path)

    html_path = Path("Rangliste.html")
    write_html_rangliste(rows, gs_results, ko_results, html_path, gs_live=gs_live)

    print(f"\n✓ Excel gespeichert:  {out_path}")
    print(f"✓ HTML gespeichert:   {html_path}")
    print("\nAktuelle Rangliste:")
    print(f"{'Platz':>5}  {'Name':<22}  {'Punkte':>6}")
    print("-" * 40)
    for r in rows:
        print(f"{r['Platz']:>5}  {r['Name']:<22}  {r['Gesamt']:>6}")

if __name__ == "__main__":
    main()
