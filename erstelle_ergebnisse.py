#!/usr/bin/env python3
"""
Erstellt Ergebnisse.xlsx – Vorlage zum Eintragen der echten Spielergebnisse.
Einmalig ausführen, dann nach jedem Spieltag befüllen.

Verwendung:
    python erstelle_ergebnisse.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

GRUPPENSPIELE = [
    {"id": "A1", "datum": "11.06.", "heim": "Mexiko",         "gast": "Südafrika",           "gruppe": "A"},
    {"id": "A2", "datum": "11.06.", "heim": "Südkorea",       "gast": "Tschechien",          "gruppe": "A"},
    {"id": "A3", "datum": "18.06.", "heim": "Tschechien",     "gast": "Südafrika",           "gruppe": "A"},
    {"id": "A4", "datum": "18.06.", "heim": "Mexiko",         "gast": "Südkorea",            "gruppe": "A"},
    {"id": "A5", "datum": "24.06.", "heim": "Tschechien",     "gast": "Mexiko",              "gruppe": "A"},
    {"id": "A6", "datum": "24.06.", "heim": "Südafrika",      "gast": "Südkorea",            "gruppe": "A"},
    {"id": "B1", "datum": "12.06.", "heim": "Kanada",         "gast": "Bosnien-Herzegowina", "gruppe": "B"},
    {"id": "B2", "datum": "13.06.", "heim": "Katar",          "gast": "Schweiz",             "gruppe": "B"},
    {"id": "B3", "datum": "18.06.", "heim": "Schweiz",        "gast": "Bosnien-Herzegowina", "gruppe": "B"},
    {"id": "B4", "datum": "18.06.", "heim": "Kanada",         "gast": "Katar",               "gruppe": "B"},
    {"id": "B5", "datum": "24.06.", "heim": "Schweiz",        "gast": "Kanada",              "gruppe": "B"},
    {"id": "B6", "datum": "24.06.", "heim": "Bosnien-Herzegowina", "gast": "Katar",          "gruppe": "B"},
    {"id": "C1", "datum": "13.06.", "heim": "Brasilien",      "gast": "Marokko",             "gruppe": "C"},
    {"id": "C2", "datum": "13.06.", "heim": "Haiti",          "gast": "Schottland",          "gruppe": "C"},
    {"id": "C3", "datum": "19.06.", "heim": "Schottland",     "gast": "Marokko",             "gruppe": "C"},
    {"id": "C4", "datum": "19.06.", "heim": "Brasilien",      "gast": "Haiti",               "gruppe": "C"},
    {"id": "C5", "datum": "24.06.", "heim": "Schottland",     "gast": "Brasilien",           "gruppe": "C"},
    {"id": "C6", "datum": "24.06.", "heim": "Marokko",        "gast": "Haiti",               "gruppe": "C"},
    {"id": "D1", "datum": "12.06.", "heim": "USA",            "gast": "Paraguay",            "gruppe": "D"},
    {"id": "D2", "datum": "13.06.", "heim": "Australien",     "gast": "Türkei",              "gruppe": "D"},
    {"id": "D3", "datum": "19.06.", "heim": "USA",            "gast": "Australien",          "gruppe": "D"},
    {"id": "D4", "datum": "19.06.", "heim": "Türkei",         "gast": "Paraguay",            "gruppe": "D"},
    {"id": "D5", "datum": "25.06.", "heim": "Türkei",         "gast": "USA",                 "gruppe": "D"},
    {"id": "D6", "datum": "25.06.", "heim": "Paraguay",       "gast": "Australien",          "gruppe": "D"},
    {"id": "E1", "datum": "14.06.", "heim": "Deutschland",    "gast": "Curaçao",             "gruppe": "E"},
    {"id": "E2", "datum": "14.06.", "heim": "Elfenbeinküste", "gast": "Ecuador",             "gruppe": "E"},
    {"id": "E3", "datum": "20.06.", "heim": "Deutschland",    "gast": "Elfenbeinküste",      "gruppe": "E"},
    {"id": "E4", "datum": "20.06.", "heim": "Ecuador",        "gast": "Curaçao",             "gruppe": "E"},
    {"id": "E5", "datum": "25.06.", "heim": "Curaçao",        "gast": "Elfenbeinküste",      "gruppe": "E"},
    {"id": "E6", "datum": "25.06.", "heim": "Ecuador",        "gast": "Deutschland",         "gruppe": "E"},
    {"id": "F1", "datum": "14.06.", "heim": "Niederlande",    "gast": "Japan",               "gruppe": "F"},
    {"id": "F2", "datum": "14.06.", "heim": "Schweden",       "gast": "Tunesien",            "gruppe": "F"},
    {"id": "F3", "datum": "20.06.", "heim": "Niederlande",    "gast": "Schweden",            "gruppe": "F"},
    {"id": "F4", "datum": "20.06.", "heim": "Tunesien",       "gast": "Japan",               "gruppe": "F"},
    {"id": "F5", "datum": "25.06.", "heim": "Japan",          "gast": "Schweden",            "gruppe": "F"},
    {"id": "F6", "datum": "25.06.", "heim": "Tunesien",       "gast": "Niederlande",         "gruppe": "F"},
    {"id": "G1", "datum": "15.06.", "heim": "Belgien",        "gast": "Ägypten",             "gruppe": "G"},
    {"id": "G2", "datum": "15.06.", "heim": "Iran",           "gast": "Neuseeland",          "gruppe": "G"},
    {"id": "G3", "datum": "21.06.", "heim": "Belgien",        "gast": "Iran",                "gruppe": "G"},
    {"id": "G4", "datum": "21.06.", "heim": "Neuseeland",     "gast": "Ägypten",             "gruppe": "G"},
    {"id": "G5", "datum": "26.06.", "heim": "Ägypten",        "gast": "Iran",                "gruppe": "G"},
    {"id": "G6", "datum": "26.06.", "heim": "Neuseeland",     "gast": "Belgien",             "gruppe": "G"},
    {"id": "H1", "datum": "15.06.", "heim": "Spanien",        "gast": "Kap Verde",           "gruppe": "H"},
    {"id": "H2", "datum": "15.06.", "heim": "Saudi-Arabien",  "gast": "Uruguay",             "gruppe": "H"},
    {"id": "H3", "datum": "21.06.", "heim": "Spanien",        "gast": "Saudi-Arabien",       "gruppe": "H"},
    {"id": "H4", "datum": "21.06.", "heim": "Uruguay",        "gast": "Kap Verde",           "gruppe": "H"},
    {"id": "H5", "datum": "26.06.", "heim": "Kap Verde",      "gast": "Saudi-Arabien",       "gruppe": "H"},
    {"id": "H6", "datum": "26.06.", "heim": "Uruguay",        "gast": "Spanien",             "gruppe": "H"},
    {"id": "I1", "datum": "16.06.", "heim": "Frankreich",     "gast": "Senegal",             "gruppe": "I"},
    {"id": "I2", "datum": "16.06.", "heim": "Irak",           "gast": "Norwegen",            "gruppe": "I"},
    {"id": "I3", "datum": "22.06.", "heim": "Frankreich",     "gast": "Irak",                "gruppe": "I"},
    {"id": "I4", "datum": "22.06.", "heim": "Norwegen",       "gast": "Senegal",             "gruppe": "I"},
    {"id": "I5", "datum": "26.06.", "heim": "Norwegen",       "gast": "Frankreich",          "gruppe": "I"},
    {"id": "I6", "datum": "26.06.", "heim": "Senegal",        "gast": "Irak",                "gruppe": "I"},
    {"id": "J1", "datum": "16.06.", "heim": "Argentinien",    "gast": "Algerien",            "gruppe": "J"},
    {"id": "J2", "datum": "16.06.", "heim": "Österreich",     "gast": "Jordanien",           "gruppe": "J"},
    {"id": "J3", "datum": "22.06.", "heim": "Argentinien",    "gast": "Österreich",          "gruppe": "J"},
    {"id": "J4", "datum": "22.06.", "heim": "Jordanien",      "gast": "Algerien",            "gruppe": "J"},
    {"id": "J5", "datum": "27.06.", "heim": "Algerien",       "gast": "Österreich",          "gruppe": "J"},
    {"id": "J6", "datum": "27.06.", "heim": "Jordanien",      "gast": "Argentinien",         "gruppe": "J"},
    {"id": "K1", "datum": "17.06.", "heim": "Portugal",       "gast": "DR Kongo",            "gruppe": "K"},
    {"id": "K2", "datum": "17.06.", "heim": "Usbekistan",     "gast": "Kolumbien",           "gruppe": "K"},
    {"id": "K3", "datum": "23.06.", "heim": "Portugal",       "gast": "Usbekistan",          "gruppe": "K"},
    {"id": "K4", "datum": "23.06.", "heim": "Kolumbien",      "gast": "DR Kongo",            "gruppe": "K"},
    {"id": "K5", "datum": "27.06.", "heim": "Kolumbien",      "gast": "Portugal",            "gruppe": "K"},
    {"id": "K6", "datum": "27.06.", "heim": "DR Kongo",       "gast": "Usbekistan",          "gruppe": "K"},
    {"id": "L1", "datum": "17.06.", "heim": "England",        "gast": "Kroatien",            "gruppe": "L"},
    {"id": "L2", "datum": "17.06.", "heim": "Ghana",          "gast": "Panama",              "gruppe": "L"},
    {"id": "L3", "datum": "23.06.", "heim": "England",        "gast": "Ghana",               "gruppe": "L"},
    {"id": "L4", "datum": "23.06.", "heim": "Panama",         "gast": "Kroatien",            "gruppe": "L"},
    {"id": "L5", "datum": "27.06.", "heim": "Panama",         "gast": "England",             "gruppe": "L"},
    {"id": "L6", "datum": "27.06.", "heim": "Kroatien",       "gast": "Ghana",               "gruppe": "L"},
]

ALLE_TEAMS = [
    "Mexiko", "Südafrika", "Südkorea", "Tschechien",
    "Kanada", "Bosnien-Herzegowina", "Katar", "Schweiz",
    "Brasilien", "Marokko", "Haiti", "Schottland",
    "USA", "Paraguay", "Australien", "Türkei",
    "Deutschland", "Curaçao", "Elfenbeinküste", "Ecuador",
    "Niederlande", "Japan", "Schweden", "Tunesien",
    "Belgien", "Ägypten", "Iran", "Neuseeland",
    "Spanien", "Kap Verde", "Saudi-Arabien", "Uruguay",
    "Frankreich", "Senegal", "Irak", "Norwegen",
    "Argentinien", "Algerien", "Österreich", "Jordanien",
    "Portugal", "DR Kongo", "Usbekistan", "Kolumbien",
    "England", "Kroatien", "Ghana", "Panama",
]

CLR_HEADER = "1A3A5C"
CLR_INPUT  = "FFFDE7"
CLR_LOCK   = "F5F5F5"
CLR_GRP    = "E8EFF7"

def hf():  return Font(bold=True, color="FFFFFF", name="Calibri", size=11)
def cf():  return Font(name="Calibri", size=10)
def bf():  return Font(bold=True, name="Calibri", size=10)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(c): return PatternFill("solid", fgColor=c)

def add_team_dropdown(ws, cell_ref, teams):
    formula = '"' + ','.join(teams[:30]) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(cell_ref)

def create_gruppenphase_sheet(wb):
    ws = wb.create_sheet("Gruppenphase")
    ws.sheet_view.showGridLines = False

    headers = ["Match ID", "Datum", "Heim", "Gast", "Ergebnis (H:G)"]
    widths  = [10,          10,       22,     22,     16]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = hf()
        cell.fill = fill(CLR_HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border()
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    gruppe = None
    row = 2
    for m in GRUPPENSPIELE:
        if m["gruppe"] != gruppe:
            gruppe = m["gruppe"]
            ws.cell(row, 1, f"Gruppe {gruppe}").font = Font(bold=True, name="Calibri", size=10, color=CLR_HEADER)
            ws.cell(row, 1).fill = fill(CLR_GRP)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1

        for c, v in enumerate([m["id"], m["datum"], m["heim"], m["gast"]], 1):
            cell = ws.cell(row, c, v)
            cell.font = cf()
            cell.fill = fill(CLR_LOCK)
            cell.alignment = Alignment(horizontal="center" if c != 3 else "left")
            cell.border = border()

        ergebnis_cell = ws.cell(row, 5)
        ergebnis_cell.font = bf()
        ergebnis_cell.fill = fill(CLR_INPUT)
        ergebnis_cell.alignment = Alignment(horizontal="center")
        ergebnis_cell.border = border()
        ergebnis_cell.comment = None

        row += 1

    # Hinweis
    ws.cell(row + 1, 1, "Hinweis: Ergebnis im Format H:G eintragen, z.B.  2:1  oder  0:0").font = Font(italic=True, color="888888", name="Calibri", size=9)
    ws.freeze_panes = "E2"

def create_ko_sheet(wb, runde_id, runde_name, n_teams):
    ws = wb.create_sheet(runde_id)
    ws.sheet_view.showGridLines = False

    ws.cell(1, 1, f"{runde_name} – Weiterkommer").font = hf()
    ws.cell(1, 1).fill = fill(CLR_HEADER)
    ws.merge_cells("A1:B1")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28

    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    for i in range(1, n_teams + 1):
        row = i + 1
        ws.cell(row, 1, i).font = cf()
        ws.cell(row, 1).fill = fill(CLR_LOCK)
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).border = border()

        team_cell = ws.cell(row, 2)
        team_cell.fill = fill(CLR_INPUT)
        team_cell.font = bf()
        team_cell.alignment = Alignment(horizontal="left")
        team_cell.border = border()
        add_team_dropdown(ws, f"B{row}", ALLE_TEAMS)

    ws.cell(n_teams + 3, 1, "Teamname aus Liste wählen oder direkt eingeben.").font = Font(italic=True, color="888888", name="Calibri", size=9)

def main():
    out = Path("Ergebnisse.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_gruppenphase_sheet(wb)
    create_ko_sheet(wb, "S16", "Sechzehntelfinale", 16)
    create_ko_sheet(wb, "S8",  "Achtelfinale",       8)
    create_ko_sheet(wb, "VF",  "Viertelfinale",       4)
    create_ko_sheet(wb, "HF",  "Halbfinale",          2)
    create_ko_sheet(wb, "F",   "Finale (Finalisten)", 2)
    create_ko_sheet(wb, "WM",  "Weltmeister",          1)

    wb.save(out)
    print(f"✓ {out} erstellt.")
    print("  → Gruppenspiel-Ergebnisse in Tab 'Gruppenphase' eintragen (Format: 2:1)")
    print("  → Nach Gruppenphase: Weiterkommer in S16, S8, VF, HF, F, WM eintragen")

if __name__ == "__main__":
    main()
